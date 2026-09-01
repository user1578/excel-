"""V2.2 样式 Schema、旧版 .xls 转换与相关离屏界面测试。"""

from __future__ import annotations

import os

import pytest
from openpyxl import Workbook, load_workbook
from PySide6.QtWidgets import QApplication, QFileDialog

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from app.ai.deepseek_parser import DeepSeekParser
from app.models.class_record import ClassRecord
from app.models.student import Student
from app.repositories.database import DatabaseManager
from app.services.class_export_service import ClassExportService, ExportColumn, SOURCE_CORE
from app.services.data_workspace_service import DataWorkspaceService
from app.services.legacy_excel_converter import LegacyExcelConversionError, LegacyExcelConverter
from app.services.master_data_service import MasterDataService
from app.services.template_service import TemplateService
from app.template_engine.schema import FieldSchema, SheetSchema, TemplateSchema
from app.template_engine.styles import WorkbookStyleSchema, business_blue_style
from app.ui.dialogs.class_students_dialog import ClassExportDialog
from app.ui.dialogs.style_dialog import StyleDialog
from app.ui.template_page import FieldEditor, TemplatePage
from app.ui.workbook_fill_page import WorkbookFillPage


class FakeTemplateClient:
    def request_template_json(self, _requirement):
        return '{"template_name":"AI样式表","student_related":false,"style":{"preset":"商务蓝色","header_fill_enabled":true,"header_fill_color":"1F4E78"},"sheets":[{"name":"录入","fields":[{"name":"事项","field_type":"text"}]}]}'


class _NoStyleClient:
    def request_template_json(self, _requirement):
        return '{"template_name":"AI默认表","student_related":false,"sheets":[{"name":"录入","fields":[{"name":"事项","field_type":"text"}]}]}'


def _master(tmp_path):
    database = DatabaseManager(tmp_path / "v22.db")
    database.initialize()
    return MasterDataService(database)


def test_standard_office_style_and_optional_title_are_rendered_and_persisted(tmp_path):
    master = _master(tmp_path)
    style = WorkbookStyleSchema(show_main_title=True, main_title="九月登记表", required_display="asterisk", default_column_width=16)
    schema = TemplateSchema("样式表", sheets=[SheetSchema("录入", [FieldSchema("姓名", required=True, column_width=20), FieldSchema("备注")])], style=style)
    artifact = TemplateService(master, tmp_path / "templates").create(schema)
    restored = TemplateSchema.from_json(artifact.schema_path.read_text(encoding="utf-8"))
    sheet = load_workbook(artifact.workbook_path)["录入"]
    assert restored.style == style
    assert sheet.merged_cells.ranges and sheet["A1"].value == "九月登记表"
    assert (sheet["A2"].value, sheet.freeze_panes, sheet.auto_filter.ref) == ("姓名*", "A3", "A2:B102")
    assert sheet["A2"].fill.fgColor.rgb in {"00000000", "000000"} and sheet["A2"].font.bold
    assert sheet["A3"].border.left.style == "thin" and sheet.row_dimensions[3].height == 22
    assert sheet.column_dimensions["A"].width == 20 and sheet.column_dimensions["B"].width == 16


def test_schema_without_style_and_ai_schema_use_safe_style_defaults(tmp_path):
    old = TemplateSchema.from_dict({"template_name": "旧模板", "sheets": [{"name": "录入", "fields": [{"name": "事项"}]}]})
    ai = DeepSeekParser(FakeTemplateClient()).generate_schema("生成一个登记表")
    ai_default = DeepSeekParser(_NoStyleClient()).generate_schema("生成一个登记表")
    assert old.style.preset == "标准办公表格"
    assert ai.style.preset == "商务蓝色" and ai.style.header_fill_enabled and ai_default.style.preset == "标准办公表格"


def test_custom_style_font_height_wrap_required_and_copy_are_rendered(tmp_path):
    master = _master(tmp_path)
    style = WorkbookStyleSchema(
        preset="自定义", overall_font_name="微软雅黑", overall_font_size=12, header_font_size=13, body_font_size=12,
        header_fill_enabled=True, header_fill_color="4472C4", header_font_color="FFFFFF", header_row_height=28,
        body_row_height=25, body_wrap_text=False, border_color="000000", required_display="cell_fill",
    )
    schema = TemplateSchema("自定义样式", default_rows=1, sheets=[SheetSchema("登记", [
        FieldSchema("姓名", required=True), FieldSchema("很长的登记事项字段名称用于测试自动列宽"),
    ])], style=style)
    service = TemplateService(master, tmp_path / "templates")
    artifact = service.create(schema)
    copied = service.copy(artifact.name)
    sheet = load_workbook(artifact.workbook_path)["登记"]
    assert (sheet["A1"].font.name, sheet["A1"].font.sz, sheet["A1"].fill.fgColor.rgb) == ("微软雅黑", 13, "004472C4")
    assert (sheet.row_dimensions[1].height, sheet.row_dimensions[2].height, bool(sheet["A2"].alignment.wrap_text)) == (28, 25, False)
    assert sheet["A2"].fill.fgColor.rgb == "00FFF2CC" and sheet["A2"].border.right.style == "thin"
    assert sheet.column_dimensions["B"].width > 20
    assert TemplateSchema.from_json(copied.schema_path.read_text(encoding="utf-8")).style == style


def test_required_display_header_color_and_default_style_do_not_apply_yellow_fill(tmp_path):
    master = _master(tmp_path)
    colored = TemplateSchema("表头必填", default_rows=1, sheets=[SheetSchema("录入", [FieldSchema("姓名", required=True)])], style=WorkbookStyleSchema(required_display="header_color"))
    plain = TemplateSchema("默认必填", default_rows=1, sheets=[SheetSchema("录入", [FieldSchema("姓名", required=True)])])
    service = TemplateService(master, tmp_path / "templates")
    colored_sheet = load_workbook(service.create(colored).workbook_path)["录入"]
    plain_sheet = load_workbook(service.create(plain).workbook_path)["录入"]
    assert colored_sheet["A1"].font.color.rgb == "00C00000"
    assert plain_sheet["A2"].fill.fill_type is None


def test_business_blue_class_export_reuses_renderer_and_column_width(tmp_path):
    master = _master(tmp_path)
    master.create_class(ClassRecord("测试班2401"))
    student = master.create_student(Student("测试学生", "20260001", "测试班2401"))
    output = ClassExportService(master, tmp_path / "exports").export("测试班2401", [student], [ExportColumn("姓名", SOURCE_CORE, "name", column_width=26)], "名单", business_blue_style())
    sheet = load_workbook(output)["学生名单"]
    assert (sheet["A1"].value, sheet["A2"].value, sheet["A2"].fill.fgColor.rgb, sheet.freeze_panes) == ("名单", "姓名", "001F4E78", "A3")
    assert sheet["A3"].border.bottom.style == "thin" and sheet.column_dimensions["A"].width == 26


class _FakeWorkbook:
    def __init__(self): self.closed = False
    def SaveAs(self, path, FileFormat):
        assert FileFormat == 51
        Workbook().save(path)
    def Close(self, SaveChanges=False): self.closed = True


class _FakeExcel:
    Visible = True
    DisplayAlerts = True
    def __init__(self):
        self.workbook = _FakeWorkbook(); self.Workbooks = self
    def Open(self, _path, ReadOnly=False):
        assert ReadOnly is True
        return self.workbook
    def Quit(self): pass


class _FakeCom:
    def __init__(self): self.excel = _FakeExcel()
    def DispatchEx(self, name):
        assert name == "Excel.Application"
        return self.excel


def test_xls_converter_creates_xlsx_copy_without_changing_original(tmp_path):
    original = tmp_path / "旧模板.xls"; original.write_bytes(b"not-a-real-xls")
    before = original.read_bytes()
    converted = LegacyExcelConverter(_FakeCom(), tmp_path).convert(original)
    assert converted.suffix == ".xlsx" and converted.is_file() and original.read_bytes() == before
    assert LegacyExcelConverter.is_legacy_template(tmp_path / "大写.XLS")


def test_xls_converter_reports_missing_excel_component(tmp_path):
    original = tmp_path / "旧模板.xls"; original.write_bytes(b"legacy")
    with pytest.raises(LegacyExcelConversionError, match="Microsoft Excel"):
        LegacyExcelConverter().convert(original)


def test_style_and_fill_dialogs_smoke(tmp_path, monkeypatch):
    application = QApplication.instance() or QApplication([])
    master = _master(tmp_path)
    master.create_class(ClassRecord("测试班2401"))
    student = master.create_student(Student("测试学生", "20260001", "测试班2401"))
    style = StyleDialog(WorkbookStyleSchema()); field = FieldEditor(field=FieldSchema("事项", column_width=18))
    export = ClassExportDialog(master, "测试班2401", [student])
    template = tmp_path / "填充模板.xlsx"; Workbook().save(template)
    page = WorkbookFillPage(DataWorkspaceService(), master)
    monkeypatch.setattr(QFileDialog, "getOpenFileName", lambda *_args, **_kwargs: (str(template), ""))
    page.choose_template(); application.processEvents()
    assert field.result_field().column_width == 18 and export.table.columnCount() == 6
    assert style.result_style().preset == "标准办公表格" and page.template_path == template
    for dialog in (style, field, export): dialog.close()
    page.close()
