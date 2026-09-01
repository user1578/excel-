"""V2.1 基础资料、文本解析与班级表功能测试；仅使用虚构资料。"""

from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from app.models.table_dataset import Provenance, TableDataset, TableRow
from app.repositories.database import DatabaseManager
from app.services.class_export_service import (
    SOURCE_AUTO, SOURCE_BLANK, SOURCE_CORE, SOURCE_EXTRA, SOURCE_FIXED,
    ClassExportService, ExportColumn,
)
from app.services.data_workspace_service import DataWorkspaceService
from app.services.master_data_import_service import MasterDataImportService
from app.services.master_data_service import MasterDataService
from app.services.text_dataset_service import TextDatasetParseError, TextDatasetService
from app.services.merge_export_service import MergeExportService
from app.models.merge_models import MergeMode, MergeResult, MergedRecord
from app.parsers.workbook_template_analyzer import WorkbookTemplateAnalyzer
from app.services.workbook_fill_service import USE_NEW_VALUE, WorkbookFillService


def service(tmp_path):
    database = DatabaseManager(tmp_path / "v21.db")
    database.initialize()
    return MasterDataService(database)


def dataset():
    columns = ["name", "student_number", "class_name", "major", "dormitory", "phone", "custom:性别"]
    return TableDataset(columns, [
        TableRow({"name": "测试学生甲", "student_number": "20260001", "class_name": "测试班2401", "major": "测试专业", "dormitory": "7栋-302", "phone": "13800000001", "custom:性别": "测试性别"}, Provenance("虚构.xlsx", "资料", 2)),
        TableRow({"name": "测试学生乙", "student_number": "20260002", "class_name": "测试班2401", "major": "测试专业", "dormitory": "无法解析寝室", "phone": "13800000002", "custom:性别": "测试性别"}, Provenance("虚构.xlsx", "资料", 3)),
    ], "虚构.xlsx", "资料", 1, column_labels={"name": "姓名", "student_number": "学号", "class_name": "班级", "major": "专业", "dormitory": "寝室", "phone": "联系电话", "custom:性别": "性别"}, custom_fields={"custom:性别"})


def test_master_import_creates_core_data_extra_fields_and_is_idempotent(tmp_path):
    master = service(tmp_path)
    importer = MasterDataImportService(master.students.database)
    preview = importer.preview(dataset())
    assert (preview.new_classes, preview.new_dormitories, preview.new_students, preview.extra_field_count) == (1, 1, 2, 1)
    assert preview.pending_dormitories == ["无法解析寝室"]
    importer.apply(dataset())
    importer.apply(dataset())
    assert master.count_classes() == 1 and master.count_dormitories() == 1 and master.count_students() == 2
    student = master.get_student_by_number("20260001")
    assert master.get_student_extra_fields(student.id)["性别"] == {"name": "性别", "value": "测试性别"}


def test_existing_student_only_fills_empty_values_and_does_not_overwrite_conflict(tmp_path):
    master = service(tmp_path)
    importer = MasterDataImportService(master.students.database)
    importer.apply(dataset())
    changed = dataset()
    changed.rows[0].values["phone"] = "13900000001"
    preview = importer.preview(changed)
    assert "20260001" in preview.student_conflicts
    importer.apply(changed)
    assert master.get_student_by_number("20260001").phone == "13800000001"


def test_extra_fields_add_update_delete_and_core_data_remains_separate(tmp_path):
    master = service(tmp_path)
    MasterDataImportService(master.students.database).apply(dataset())
    student = master.get_student_by_number("20260001")
    master.set_student_extra_field(student.id, "自定义名称", "值 A")
    master.set_student_extra_field(student.id, "自定义名称", "值 B")
    fields = master.get_student_extra_fields(student.id)
    assert fields["自定义名称"]["value"] == "值 B" and "自定义名称" not in student.__dict__
    assert master.delete_student_extra_field(student.id, "自定义名称")
    assert "自定义名称" not in master.get_student_extra_fields(student.id)


def test_master_import_transaction_rolls_back_when_later_write_fails(tmp_path, monkeypatch):
    master = service(tmp_path)
    importer = MasterDataImportService(master.students.database)
    original = importer._parse_dormitory
    calls = {"count": 0}
    def broken(values):
        calls["count"] += 1
        if calls["count"] > 2:
            raise RuntimeError("模拟导入异常")
        return original(values)
    monkeypatch.setattr(importer, "_parse_dormitory", broken)
    with pytest.raises(RuntimeError):
        importer.apply(dataset())
    assert master.count_classes() == master.count_dormitories() == master.count_students() == 0


def test_text_parser_supports_tsv_key_value_encodings_and_rejects_free_form(tmp_path):
    parser = TextDatasetService()
    tsv = parser.parse_text("姓名\t学号\t班级\n测试学生甲\t20260001\t测试班2401")
    key_value = parser.parse_text("姓名：测试学生甲\n学号=20260001\n\n姓名:测试学生乙\n学号:20260002")
    path = tmp_path / "虚构.txt"; path.write_bytes("姓名：测试学生甲\n学号：20260001".encode("gb18030"))
    assert (len(tsv.rows), tsv.columns[:2], len(key_value.rows), parser.parse_file(path).rows[0].values["student_number"]) == (1, ["name", "student_number"], 2, "20260001")
    try:
        parser.parse_text("测试学生甲，测试班2401，参加比赛")
    except TextDatasetParseError:
        pass
    else:
        raise AssertionError("自由文本必须明确提示需要 AI 结构化解析")


@pytest.mark.parametrize("encoding", ["utf-8", "utf-8-sig", "gb18030"])
def test_txt_common_encodings_and_single_record_are_supported(tmp_path, encoding):
    path = tmp_path / f"{encoding}.txt"
    path.write_bytes("姓名：测试学生甲\n学号：20260001\n参赛：是".encode(encoding))
    dataset_value = TextDatasetService().parse_file(path)
    assert len(dataset_value.rows) == 1 and dataset_value.rows[0].values["custom:参赛"] == "是"


def test_text_ai_json_is_only_a_dataset_and_invalid_schema_is_rejected():
    parser = TextDatasetService()
    result = parser.parse_ai_json('{"columns":["姓名","学号"],"rows":[{"姓名":"测试学生甲","学号":"20260001"}]}')
    assert result.rows[0].values["name"] == "测试学生甲"
    with pytest.raises(TextDatasetParseError):
        parser.parse_ai_json('{"rows": []}')


def test_class_export_all_column_types_schemes_and_dataset_link(tmp_path):
    master = service(tmp_path)
    MasterDataImportService(master.students.database).apply(dataset())
    students = master.list_students_by_class("测试班2401")
    exporter = ClassExportService(master, tmp_path / "exports")
    columns = [
        ExportColumn("序号", SOURCE_AUTO), ExportColumn("姓名", SOURCE_CORE, "name"),
        ExportColumn("性别", SOURCE_EXTRA, "性别"), ExportColumn("参赛", SOURCE_BLANK), ExportColumn("学院", SOURCE_FIXED, fixed_value="测试学院"),
    ]
    output = exporter.export("测试班2401", students, columns, "测试班2401报名表")
    sheet = load_workbook(output)["学生名单"]
    assert (sheet["A1"].value, sheet["A3"].value, sheet["B3"].value, sheet["C3"].value, sheet["D3"].value, sheet["E3"].value) == ("测试班2401报名表", 1, "测试学生甲", "测试性别", None, "测试学院")
    scheme = exporter.save_scheme("测试报名名单", "测试标题", columns)
    assert exporter.list_schemes()[0].name == "测试报名名单" and exporter.delete_scheme(scheme.id)
    workspace = DataWorkspaceService(); workspace.set_dataset(exporter.students_dataset(students))
    assert workspace.current_merge_result is None and "extra:性别" in workspace.current_dataset.columns


def test_class_export_without_title_unique_paths_and_safe_values(tmp_path):
    master = service(tmp_path)
    MasterDataImportService(master.students.database).apply(dataset())
    students = master.list_students_by_class("测试班2401")
    exporter = ClassExportService(master, tmp_path / "exports")
    columns = [ExportColumn("姓名", SOURCE_CORE, "name"), ExportColumn("备注", SOURCE_FIXED, fixed_value="=不执行")]
    first = exporter.export("测试班2401", students, columns)
    second = exporter.export("测试班2401", students, columns)
    sheet = load_workbook(first)["学生名单"]
    assert first != second and sheet["A1"].value == "姓名" and sheet["B2"].value == "'=不执行" and sheet.freeze_panes == "A2"


def test_unmatched_and_unresolved_records_are_both_counted_once():
    result = MergeResult(MergeMode.STUDENT, ["name"], {"name": "姓名"}, [
        MergedRecord({"name": "测试学生甲"}, match_status="unmatched"),
        MergedRecord({"name": "测试学生乙"}, match_status="unresolved"),
    ], unresolved_record_indexes=[1])
    assert MergeExportService._unlinked_count(result) == 2


def test_text_and_selected_class_dataset_fill_existing_template_without_changing_it(tmp_path):
    template = tmp_path / "虚构学校模板.xlsx"
    workbook = Workbook(); sheet = workbook.active; sheet.title = "报名"
    sheet.append(["姓名", "学号", "班级", "性别"]); workbook.save(template)
    before = template.read_bytes()
    analyzer = WorkbookTemplateAnalyzer().analyze(template, "报名", 1)
    text_dataset = TextDatasetService().parse_text("姓名：测试学生甲\n学号：20260001\n班级：测试班2401\n\n姓名：测试学生乙\n学号：20260002\n班级：测试班2401")
    text_result = WorkbookFillService(tmp_path / "text_exports").fill(analyzer, text_dataset, {"姓名": "name", "学号": "student_number", "班级": "class_name"}, USE_NEW_VALUE)
    assert template.read_bytes() == before and load_workbook(text_result.output_path)["报名"]["A2"].value == "测试学生甲"
    master = service(tmp_path); MasterDataImportService(master.students.database).apply(dataset())
    class_dataset = ClassExportService(master).students_dataset(master.list_students_by_class("测试班2401"))
    class_result = WorkbookFillService(tmp_path / "class_exports").fill(analyzer, class_dataset, {"姓名": "name", "学号": "student_number", "班级": "class_name", "性别": "extra:性别"}, USE_NEW_VALUE)
    assert load_workbook(class_result.output_path)["报名"]["D2"].value == "测试性别"


def test_master_import_reports_invalid_duplicate_rows_and_never_silently_skips(tmp_path):
    master = service(tmp_path)
    source = dataset()
    source.rows.append(TableRow({"name": "", "student_number": "20260003", "class_name": "测试班2401"}, Provenance("缺失.xlsx", "资料", 9)))
    source.rows.append(TableRow({"name": "冲突学生", "student_number": "20260001", "class_name": "其他班2401"}, Provenance("重复.xlsx", "资料", 10)))
    preview = MasterDataImportService(master.students.database).preview(source)
    assert preview.skipped_count == 2
    assert {(item.source_file, item.source_sheet, item.source_row) for item in preview.skipped_records} == {("缺失.xlsx", "资料", 9), ("重复.xlsx", "资料", 10)}
    assert "20260001" in preview.student_conflicts and "缺少必填字段" in preview.invalid_records[0].reason
    MasterDataImportService(master.students.database).apply(source)
    assert master.count_students() == 2


def test_dormitory_correction_is_explicit_and_skip_does_not_write_unconfirmed_value(tmp_path):
    master = service(tmp_path)
    source = dataset()
    importer = MasterDataImportService(master.students.database)
    correction = {"无法解析寝室": ("测试7栋", "303")}
    preview = importer.preview(source, correction)
    assert not preview.pending_dormitories
    importer.apply(source, correction)
    assert master.get_student_by_number("20260002").dormitory == "测试7栋-303"
    other = service(tmp_path / "skip")
    MasterDataImportService(other.students.database).apply(dataset(), {"无法解析寝室": None})
    assert other.get_student_by_number("20260002").dormitory is None


def test_import_only_preserves_custom_student_fields_not_business_columns(tmp_path):
    master = service(tmp_path)
    source = dataset()
    source.columns.extend(["sequence", "date", "course", "status", "building", "room_number", "custom:政治面貌"])
    source.custom_fields.update({"sequence", "date", "course", "status", "building", "room_number", "custom:政治面貌"})
    source.column_labels.update({"sequence": "序号", "date": "日期", "course": "课程", "status": "状态", "building": "楼栋", "room_number": "寝室号", "custom:政治面貌": "政治面貌"})
    for row in source.rows:
        row.values.update({"sequence": "1", "date": "2026-09-01", "course": "测试课程", "status": "正常", "building": "测试7栋", "room_number": "302", "custom:政治面貌": "测试身份"})
    MasterDataImportService(master.students.database).apply(source)
    fields = master.get_student_extra_fields(master.get_student_by_number("20260001").id)
    assert set(fields) == {"性别", "政治面貌"}


def test_renaming_extra_field_replaces_old_key_in_one_change(tmp_path):
    master = service(tmp_path)
    MasterDataImportService(master.students.database).apply(dataset())
    student = master.get_student_by_number("20260001")
    master.set_student_extra_field(student.id, "政治面貌", "测试身份")
    master.apply_student_extra_field_changes(student.id, [("政治状态", "测试身份", "政治面貌")])
    fields = master.get_student_extra_fields(student.id)
    assert "政治面貌" not in fields and fields["政治状态"] == {"name": "政治状态", "value": "测试身份"}


def test_custom_text_columns_with_same_normalized_key_keep_both_values():
    result = TextDatasetService().parse_text("父亲-电话\t父亲 电话\n111\t222")
    assert result.columns == ["custom:父亲电话", "custom:父亲电话__2"]
    assert result.rows[0].values == {"custom:父亲电话": "111", "custom:父亲电话__2": "222"}
