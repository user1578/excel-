"""TemplateSchema、本地模板工作簿、管理、DeepSeek mock 与离屏 UI 测试。"""

from __future__ import annotations

import os
import socket
import urllib.error
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PySide6.QtWidgets import QApplication, QMessageBox, QPushButton

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from app.ai.deepseek_client import DeepSeekClient, DeepSeekClientError, DeepSeekConfig
from app.ai.deepseek_parser import DeepSeekParser
from app.models.class_record import ClassRecord
from app.models.dormitory import Dormitory
from app.models.field_mapping import StandardField
from app.models.student import Student
from app.repositories.database import DatabaseManager
from app.services.master_data_service import MasterDataService
from app.services.template_service import TemplateService
from app.template_engine.schema import FieldSchema, SchemaValidationError, SheetSchema, TemplateSchema
from app.ui.template_page import TemplatePage
from app.ui.settings_page import SettingsPage


@pytest.fixture
def template_setup(tmp_path):
    database = DatabaseManager(tmp_path / "data" / "database.db")
    database.initialize()
    master = MasterDataService(database)
    master.create_class(ClassRecord("物联网2401"))
    master.create_class(ClassRecord("软件2401"))
    master.create_dormitory(Dormitory("南3", "402", "南3-402"))
    master.create_student(Student("张三", "20260001", "物联网2401"))
    return TemplateService(master, tmp_path / "templates"), tmp_path


def classroom_schema() -> TemplateSchema:
    return TemplateSchema("课堂查课表", True, "课堂检查", 3, [SheetSchema("数据录入", [
        FieldSchema("姓名", "name", True, standard_field="name", allow_blank=False),
        FieldSchema("学号", "student_number", True, standard_field="student_number", allow_blank=False),
        FieldSchema("班级", "class_name", True, data_source="classes", standard_field="class_name", allow_blank=False),
        FieldSchema("日期", "date", True, standard_field="date"), FieldSchema("课程", "text", True, standard_field="course"),
        FieldSchema("应到", "integer", standard_field="expected"), FieldSchema("实到", "integer", standard_field="actual"),
        FieldSchema("到课率", "formula", formula="{实到}/{应到}"), FieldSchema("状态", "select", options=["正常", "迟到", "缺勤", "请假"], standard_field="status"),
        FieldSchema("备注", "text", standard_field="remark"),
    ])])


def test_schema_json_round_trip_and_core_defaults():
    schema = classroom_schema()
    restored = TemplateSchema.from_json(schema.to_json())
    assert restored == schema
    missing = TemplateSchema("学生名单", True, sheets=[SheetSchema("录入", [FieldSchema("联系电话", standard_field="phone")])]).ensure_student_core_fields()
    assert [field.standard_field for field in missing.sheets[0].fields][:3] == ["name", "student_number", "class_name"]
    non_student = TemplateSchema("物品表", False, sheets=[SheetSchema("物品", [FieldSchema("物品名称")])]).ensure_student_core_fields()
    assert [field.name for field in non_student.sheets[0].fields] == ["物品名称"]
    reordered = TemplateSchema("早读", True, sheets=[SheetSchema("录入", [FieldSchema("日期", "date"), FieldSchema("班级", "class_name", standard_field="class_name"), FieldSchema("姓名", "name", standard_field="name"), FieldSchema("学号", "student_number", standard_field="student_number")])]).ensure_student_core_fields()
    assert [field.name for field in reordered.sheets[0].fields][:3] == ["姓名", "学号", "班级"]


@pytest.mark.parametrize("schema, message", [
    (TemplateSchema("重复", sheets=[SheetSchema("录入", [FieldSchema("姓名"), FieldSchema("姓名")])]), "字段名重复"),
    (TemplateSchema("工作表", sheets=[SheetSchema("非法/名称", [FieldSchema("字段")])]), "工作表名称不合法"),
    (TemplateSchema("公式", sheets=[SheetSchema("录入", [FieldSchema("到课率", "formula", formula="{不存在}/{应到}"), FieldSchema("应到", "integer")])]), "公式引用的字段不存在"),
    (TemplateSchema("下拉", sheets=[SheetSchema("录入", [FieldSchema("状态", "select")])]), "必须配置选项或数据源"),
    (TemplateSchema("安全", sheets=[SheetSchema("录入", [FieldSchema("备注", default_value="=HYPERLINK(\"x\")")])]), "默认值不能包含 Excel 公式"),
])
def test_schema_validation_rejects_invalid_configs(schema, message):
    with pytest.raises(SchemaValidationError, match=message): schema.validate()


def test_generator_creates_validation_formats_formula_and_hidden_data(template_setup):
    service, _tmp_path = template_setup
    artifact = service.create(classroom_schema())
    workbook = load_workbook(artifact.workbook_path)
    sheet = workbook["数据录入"]
    assert workbook["_系统数据"].sheet_state == "hidden"
    assert workbook["使用说明"]["A1"].value == "模板名称"
    assert sheet.freeze_panes is None and sheet.auto_filter.ref is None
    assert sheet["D2"].number_format == "yyyy-mm-dd"
    assert sheet["H2"].value == "=IFERROR(G2/F2,0)"
    assert len(sheet.data_validations.dataValidation) == 2
    assert sheet["A1"].comment.text == "必填字段"


def test_formula_survives_field_order_change(template_setup):
    service, _tmp_path = template_setup
    schema = TemplateSchema("比例", sheets=[SheetSchema("录入", [FieldSchema("比例", "formula", formula="{实到}/{应到}"), FieldSchema("实到", "integer"), FieldSchema("应到", "integer")])])
    artifact = service.create(schema)
    assert load_workbook(artifact.workbook_path, data_only=False)["录入"]["A2"].value == "=IFERROR(B2/C2,0)"


def test_template_manager_load_copy_delete_and_does_not_overwrite(template_setup):
    service, _tmp_path = template_setup
    first = service.create(classroom_schema())
    second = service.create(classroom_schema())
    assert first.directory != second.directory
    assert service.load(first.name).template_name == "课堂查课表"
    copied = service.copy(first.name)
    assert copied.directory.exists() and copied.schema_path.exists()
    service.delete(copied.name)
    assert not copied.directory.exists()


class FakeClient:
    def __init__(self, response=None, error=None): self.response, self.error = response, error
    def request_template_json(self, _requirement):
        if self.error: raise self.error
        return self.response


def test_ai_parser_accepts_json_and_auto_adds_student_core_fields():
    parser = DeepSeekParser(FakeClient('{"template_name":"早读迟到表","student_related":true,"default_rows":10,"sheets":[{"name":"录入","fields":[{"name":"迟到","field_type":"integer","standard_field":"late"}]}]}'))
    schema = parser.generate_schema("做早读迟到表")
    assert [field.standard_field for field in schema.sheets[0].fields][:3] == ["name", "student_number", "class_name"]


def test_ai_parser_accepts_limited_markdown_and_field_type_aliases():
    parser = DeepSeekParser(FakeClient('方案如下：```json\n{"template_name":"别名表","student_related":false,"sheets":[{"name":"录入","fields":[{"name":"日期","field_type":"日期","standard_field":"日期"},{"name":"事项","field_type":"文本"}]}]}\n```'))
    schema = parser.generate_schema("需求")
    assert [(field.field_type, field.standard_field) for field in schema.sheets[0].fields] == [("date", "date"), ("text", None)]


def test_ai_parser_treats_null_optional_lists_as_empty():
    schema = DeepSeekParser(FakeClient('{"template_name":"空选项","student_related":false,"sheets":[{"name":"录入","fields":[{"name":"事项","field_type":"text","options":null}]}]}')).generate_schema("需求")
    assert schema.sheets[0].fields[0].options == []


def test_ai_parser_removes_only_formula_leading_equals():
    schema = DeepSeekParser(FakeClient('{"template_name":"比例","student_related":false,"sheets":[{"name":"录入","fields":[{"name":"实到","field_type":"整数"},{"name":"应到","field_type":"整数"},{"name":"比例","field_type":"计算字段","formula":"={实到}/{应到}"}]}]}')).generate_schema("需求")
    assert schema.sheets[0].fields[-1].formula == "{实到}/{应到}"


def test_ai_parser_standardizes_known_fields_for_local_generation():
    schema = DeepSeekParser(FakeClient('{"template_name":"查课","student_related":true,"sheets":[{"name":"录入","fields":[{"name":"姓名","field_type":"text"},{"name":"班级","field_type":"text"},{"name":"寝室","field_type":"text"},{"name":"检查日期","field_type":"text"}]}]}')).generate_schema("需求")
    fields = {field.name: field for field in schema.sheets[0].fields}
    assert (fields["姓名"].field_type, fields["姓名"].required) == ("name", True)
    assert (fields["班级"].data_source, fields["寝室"].data_source, fields["检查日期"].field_type) == ("classes", "dormitories", "date")


def test_ai_parser_treats_known_business_problem_field_as_custom():
    schema = DeepSeekParser(FakeClient('{"template_name":"查寝","student_related":false,"sheets":[{"name":"录入","fields":[{"name":"存在问题","field_type":"text","standard_field":"问题"}]}]}')).generate_schema("需求")
    assert schema.sheets[0].fields[0].standard_field is None


def test_non_student_schema_does_not_keep_mislabelled_core_field():
    schema = DeepSeekParser(FakeClient('{"template_name":"物品","student_related":false,"sheets":[{"name":"录入","fields":[{"name":"负责人","field_type":"姓名","standard_field":"name"}]}]}')).generate_schema("需求")
    field = schema.sheets[0].fields[0]
    assert (field.field_type, field.standard_field) == ("text", None)


@pytest.mark.parametrize("response", ["不是 JSON", '{"template_name":"错误","sheets":[]}'])
def test_ai_parser_rejects_non_json_or_invalid_schema(response):
    with pytest.raises(SchemaValidationError): DeepSeekParser(FakeClient(response)).generate_schema("任意需求")


@pytest.mark.parametrize("error", [
    DeepSeekClientError("DeepSeek 身份验证失败，请检查 API Key。"), DeepSeekClientError("DeepSeek 请求过于频繁，请稍后再试。"),
    DeepSeekClientError("无法连接 DeepSeek 或请求超时，可继续使用手动创建。"),
])
def test_ai_parser_surfaces_safe_client_errors(error):
    with pytest.raises(DeepSeekClientError): DeepSeekParser(FakeClient(error=error)).generate_schema("任意需求")


def test_deepseek_client_without_key_falls_back_safely(tmp_path):
    env = tmp_path / ".env"
    env.write_text("DEEPSEEK_ENABLED=true\nDEEPSEEK_API_KEY=\nDEEPSEEK_BASE_URL=https://test\n", encoding="utf-8")
    with pytest.raises(DeepSeekClientError, match="API Key 未配置"):
        DeepSeekClient(env).request_template_json("需求")


class FakeResponse:
    def __init__(self, content='{"choices":[{"message":{"content":"OK"}}]}'): self.content = content.encode("utf-8")
    def __enter__(self): return self
    def __exit__(self, *_args): return False
    def read(self): return self.content


def test_connection_success_and_config_reloads_without_exposing_key(tmp_path, monkeypatch):
    env = tmp_path / ".env"
    DeepSeekConfig(True, "private-test-key", "first-model", "https://first").save(env)
    calls = []
    def fake_open(request, **_kwargs):
        calls.append(request); return FakeResponse()
    monkeypatch.setattr("urllib.request.urlopen", fake_open)
    client = DeepSeekClient(env)
    client.test_connection()
    DeepSeekConfig(True, "private-test-key", "second-model", "https://second").save(env)
    client.test_connection()
    assert calls[1].full_url == "https://second/chat/completions"
    assert b"second-model" in calls[1].data
    assert "private-test-key" not in str(DeepSeekClientError("网络失败"))


def test_env_example_is_root_level_and_gitignore_excludes_only_real_env():
    root = Path(__file__).resolve().parents[1]
    assert (root / ".env.example").is_file()
    ignore = (root / ".gitignore").read_text(encoding="utf-8")
    assert ".env" in ignore and ".env.example" not in ignore


def test_settings_page_has_connection_test_button_without_real_request(tmp_path, monkeypatch):
    class Client:
        def __init__(self): self.called = False
        def test_connection(self): self.called = True
    client = Client(); application = QApplication.instance() or QApplication([])
    page = SettingsPage(tmp_path / ".env", client=client)
    monkeypatch.setattr(QMessageBox, "information", lambda *_args: None)
    page.test_connection(); application.processEvents()
    assert client.called
    assert "测试连接" in [button.text() for button in page.findChildren(QPushButton)]
    page.close()


def test_connection_authentication_failure_is_safe(tmp_path, monkeypatch):
    env = tmp_path / ".env"; DeepSeekConfig(True, "hidden-key", "model", "https://test").save(env)
    error = urllib.error.HTTPError("https://test", 401, "", {}, None)
    monkeypatch.setattr("urllib.request.urlopen", lambda *_args, **_kwargs: (_ for _ in ()).throw(error))
    with pytest.raises(DeepSeekClientError, match="身份验证失败") as raised:
        DeepSeekClient(env).test_connection()
    assert "hidden-key" not in str(raised.value)


@pytest.mark.parametrize("error, expected", [
    (urllib.error.HTTPError("https://test", 401, "", {}, None), "身份验证失败"),
    (urllib.error.HTTPError("https://test", 429, "", {}, None), "请求过于频繁"),
    (socket.timeout(), "无法连接"),
    (urllib.error.URLError("offline"), "无法连接"),
])
def test_deepseek_client_maps_http_timeout_and_network_errors(tmp_path, monkeypatch, error, expected):
    env = tmp_path / ".env"; env.write_text("DEEPSEEK_ENABLED=true\nDEEPSEEK_API_KEY=test-key\nDEEPSEEK_MODEL=test\nDEEPSEEK_BASE_URL=https://test\n", encoding="utf-8")
    monkeypatch.setattr("urllib.request.urlopen", lambda *_args, **_kwargs: (_ for _ in ()).throw(error))
    with pytest.raises(DeepSeekClientError, match=expected): DeepSeekClient(env).request_template_json("需求")


def test_template_page_smoke_manual_generation_and_ai_area(template_setup, monkeypatch):
    service, tmp_path = template_setup
    application = QApplication.instance() or QApplication([])
    parser = DeepSeekParser(FakeClient('{"template_name":"AI模板","student_related":false,"default_rows":2,"sheets":[{"name":"录入","fields":[{"name":"事项","field_type":"text"}]}]}'))
    page = TemplatePage(service, parser); page.show(); application.processEvents()
    page.name.setText("手动学生表"); page.student_related.setChecked(True)
    assert [field.name for field in page.fields][:3] == ["姓名", "学号", "班级"]
    page.fields.append(FieldSchema("备注")); page._render_fields()
    page.style = replace(page.style, title_mode="none")
    monkeypatch.setattr(QMessageBox, "information", lambda *_args: None)
    page.generate_template()
    assert page.template_table.rowCount() == 1
    page.template_table.selectRow(0); page.open_selected()
    assert page.name.text() == "手动学生表"
    page.ai_input.setPlainText("做一个简单事项表"); page.generate_ai_schema()
    assert page.name.text() == "AI模板"
    assert page.ai_status.text() == "AI 方案已生成，请确认后再生成 Excel 模板。"
    assert page.ai_generate_button.isEnabled()
    page.close()
