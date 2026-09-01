"""表格填充服务测试；模板和数据均在临时目录生成。"""

from __future__ import annotations

import hashlib

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.models.table_dataset import Provenance, TableDataset, TableRow
from app.parsers.workbook_template_analyzer import WorkbookTemplateAnalyzer
from app.services.workbook_fill_service import (
    AUTO_SEQUENCE,
    KEEP_EXISTING,
    SKIP_CONFLICTING_ROW,
    USE_NEW_VALUE,
    MergedCellWriteError,
    WorkbookFillService,
)


def source_dataset():
    columns = ["name", "student_number", "class_name", "phone", "custom:note"]
    return TableDataset(
        columns,
        [
            TableRow({"name": "测试学生甲", "student_number": "20260001", "class_name": "测试班2401", "phone": "13800000001", "custom:note": "=1+1"}, Provenance("来源.xlsx", "资料", 2)),
            TableRow({"name": "测试学生乙", "student_number": "20260002", "class_name": "测试班2401", "phone": "13800000002", "custom:note": "普通备注"}, Provenance("来源.xlsx", "资料", 3)),
        ],
        "来源.xlsx", "资料", 1,
        column_labels={"name": "姓名", "student_number": "学号", "class_name": "班级", "phone": "手机号", "custom:note": "备注"},
        custom_fields={"custom:note"},
    )


def make_template(path, existing_name=""):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "汇总表"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "虚构学院统计模板"
    sheet.append([])
    sheet.append(["序号", "学生姓名", "学籍号", "所在行政班", "联系电话", "备注", "相对公式", "混合公式"])
    sheet.append([1, existing_name, "", "", "", "", "=B4&\"-\"&C4", "=$B4&C$3"])
    style = Font(bold=True, color="112233")
    for cell in sheet[4]:
        cell.font = style
        cell.fill = PatternFill("solid", fgColor="DDEEFF")
        cell.border = Border(bottom=Side(style="thin", color="123456"))
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = "@"
    sheet.row_dimensions[4].height = 25
    sheet.column_dimensions["B"].width = 24
    sheet.freeze_panes = "A4"
    workbook.save(path)


def digest(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_fill_creates_new_file_preserves_template_and_copies_style_and_formulas(tmp_path):
    template = tmp_path / "学院统计模板.xlsx"
    make_template(template)
    before = digest(template)
    analysis = WorkbookTemplateAnalyzer().analyze(template, "汇总表", 3)
    service = WorkbookFillService(tmp_path / "exports")
    mappings = service.default_mappings(analysis, source_dataset())
    assert mappings == {"序号": AUTO_SEQUENCE, "学生姓名": "name", "学籍号": "student_number", "所在行政班": "class_name", "联系电话": "phone", "备注": "custom:note"}
    result = service.fill(analysis, source_dataset(), mappings, USE_NEW_VALUE)

    assert digest(template) == before
    assert result.output_path != template and result.output_path.exists()
    workbook = load_workbook(result.output_path, data_only=False)
    sheet = workbook["汇总表"]
    assert sheet["B4"].value == "测试学生甲"
    assert sheet["F4"].value == "'=1+1"
    assert sheet["G5"].value == '=B5&"-"&C5'
    assert sheet["H5"].value == "=$B5&C$3"
    assert sheet["B5"].font.bold is True and sheet["B5"].font.color.rgb == "00112233"
    assert sheet["B5"].fill.fgColor.rgb == "00DDEEFF"
    assert sheet["B5"].border.bottom.style == "thin"
    assert sheet["B5"].alignment.horizontal == "center"
    assert sheet["B5"].number_format == "@"
    assert sheet.row_dimensions[5].height == 25
    assert sheet.column_dimensions["B"].width == 24
    assert sheet.freeze_panes == "A4"
    assert "A1:F1" in {str(item) for item in sheet.merged_cells.ranges}


@pytest.mark.parametrize("strategy, expected, skipped, preserved", [
    (KEEP_EXISTING, "模板原值", 0, 1),
    (USE_NEW_VALUE, "测试学生甲", 0, 0),
    (SKIP_CONFLICTING_ROW, "模板原值", 1, 0),
])
def test_existing_value_strategies_are_explicit(tmp_path, strategy, expected, skipped, preserved):
    template = tmp_path / f"策略_{strategy}.xlsx"
    make_template(template, "模板原值")
    analysis = WorkbookTemplateAnalyzer().analyze(template, "汇总表", 3)
    service = WorkbookFillService(tmp_path / "exports")
    result = service.fill(analysis, source_dataset(), {"学生姓名": "name"}, strategy)
    sheet = load_workbook(result.output_path)["汇总表"]
    assert sheet["B4"].value == expected
    assert result.skipped_rows == skipped
    assert result.preserved_cells == preserved


def test_merged_non_anchor_is_rejected_without_changing_template(tmp_path):
    template = tmp_path / "合并单元格模板.xlsx"
    make_template(template)
    workbook = load_workbook(template)
    sheet = workbook["汇总表"]
    sheet.merge_cells("C4:D4")
    workbook.save(template)
    before = digest(template)
    analysis = WorkbookTemplateAnalyzer().analyze(template, "汇总表", 3)
    service = WorkbookFillService(tmp_path / "exports")
    with pytest.raises(MergedCellWriteError, match="非左上角"):
        service.fill(analysis, source_dataset(), {"所在行政班": "class_name"}, USE_NEW_VALUE)
    assert digest(template) == before


def test_preview_reports_existing_values_and_unique_outputs(tmp_path):
    template = tmp_path / "预览模板.xlsx"
    make_template(template, "模板原值")
    analysis = WorkbookTemplateAnalyzer().analyze(template, "汇总表", 3)
    service = WorkbookFillService(tmp_path / "exports")
    preview = service.preview(analysis, source_dataset(), {"学生姓名": "name"})
    assert preview.row_count == 2 and preview.existing_value_conflicts == 1
    first = service.fill(analysis, source_dataset(), {"学生姓名": "name"}, USE_NEW_VALUE)
    second = service.fill(analysis, source_dataset(), {"学生姓名": "name"}, USE_NEW_VALUE)
    assert first.output_path != second.output_path
