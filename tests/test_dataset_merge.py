"""资料汇总服务测试；文件和人员均为虚构数据。"""

from __future__ import annotations

from openpyxl import Workbook, load_workbook

from app.models.merge_models import ConflictResolution, MergeMode
from app.models.table_dataset import Provenance, TableDataset, TableRow
from app.services.dataset_merge_service import DatasetMergeService
from app.services.merge_export_service import MergeExportService
from app.services.table_analysis_service import TableAnalysisService
from app.services.data_workspace_service import DataWorkspaceService


def dataset(name, columns, rows, labels=None, sheet="资料"):
    return TableDataset(
        columns,
        [TableRow(values, Provenance(name, sheet, index + 2)) for index, values in enumerate(rows)],
        name,
        sheet,
        1,
        column_labels=labels or {key: key for key in columns},
        custom_fields={key for key in columns if key.startswith("custom:")},
    )


def test_vertical_merge_uses_column_union_without_order_misalignment():
    service = DatasetMergeService()
    left = dataset("班一.xlsx", ["name", "student_number", "phone"], [{"name": "测试学生甲", "student_number": "20260001", "phone": "13800000001"}])
    right = dataset("班二.xlsx", ["student_number", "name", "custom:dorm"], [{"student_number": "20260002", "name": "测试学生乙", "custom:dorm": "A-101"}], {"student_number": "学号", "name": "姓名", "custom:dorm": "寝室"})
    result = service.merge_vertical([left, right])
    assert result.columns == ["name", "student_number", "phone", "custom:dorm"]
    assert result.records[1].values["name"] == "测试学生乙"
    assert result.records[1].values["phone"] == ""
    assert result.records[1].provenance["custom:dorm"].source_file == "班二.xlsx"


def test_student_merge_prioritizes_number_then_name_and_class_then_safe_unique_name():
    service = DatasetMergeService()
    a = dataset("基础.xlsx", ["name", "student_number", "class_name", "phone"], [
        {"name": "测试学生甲", "student_number": "20260001", "class_name": "测试班2401", "phone": "13800000001"},
        {"name": "测试学生乙", "class_name": "测试班2402"},
        {"name": "测试学生丙"},
    ])
    b = dataset("补充.xlsx", ["name", "student_number", "class_name", "custom:award"], [
        {"name": "别名甲", "student_number": "20260001", "class_name": "测试班2401", "custom:award": "一等奖"},
        {"name": "测试学生乙", "class_name": "测试班2402", "custom:award": "二等奖"},
        {"name": "测试学生丙", "custom:award": "三等奖"},
    ])
    result = service.merge_by_student([a, b])
    assert len(result.records) == 3
    assert result.records[0].values["custom:award"] == "一等奖"
    assert result.records[1].values["custom:award"] == "二等奖"
    assert result.records[2].values["custom:award"] == "三等奖"


def test_same_name_different_class_never_auto_merges():
    service = DatasetMergeService()
    result = service.merge_by_student([
        dataset("一.xlsx", ["name", "class_name"], [{"name": "测试重名", "class_name": "测试班2401"}]),
        dataset("二.xlsx", ["name", "class_name"], [{"name": "测试重名", "class_name": "测试班2402"}]),
    ])
    assert len(result.records) == 2
    assert {record.values["class_name"] for record in result.records} == {"测试班2401", "测试班2402"}


def test_conflicts_never_silently_overwrite_and_can_be_resolved():
    service = DatasetMergeService()
    result = service.merge_by_student([
        dataset("一.xlsx", ["name", "student_number", "phone"], [{"name": "测试学生甲", "student_number": "20260001", "phone": "13800000001"}]),
        dataset("二.xlsx", ["name", "student_number", "phone"], [{"name": "测试学生甲", "student_number": "20260001", "phone": "13900000001"}]),
    ])
    conflict = next(item for item in result.conflicts if item.field == "phone")
    assert result.records[0].values["phone"] == "13800000001"
    assert not conflict.is_resolved
    service.resolve_conflict(result, conflict.id, ConflictResolution.USE_B)
    assert result.records[0].values["phone"] == "13900000001"
    service.resolve_conflict(result, conflict.id, ConflictResolution.MANUAL, "13700000001")
    assert result.records[0].values["phone"] == "13700000001"


def test_export_requires_explicit_permission_for_unresolved_conflicts_and_escapes_user_text(tmp_path):
    service = DatasetMergeService()
    result = service.merge_by_student([
        dataset("一.xlsx", ["name", "student_number", "custom:note"], [{"name": "测试学生甲", "student_number": "20260001", "custom:note": "=1+1"}]),
        dataset("二.xlsx", ["name", "student_number", "custom:note"], [{"name": "测试学生甲", "student_number": "20260001", "custom:note": "+bad"}]),
    ])
    exporter = MergeExportService(tmp_path / "exports")
    try:
        exporter.export(result)
        raise AssertionError("未解决冲突不得静默导出")
    except ValueError:
        pass
    path = exporter.export(result, allow_unresolved=True)
    workbook = load_workbook(path, data_only=False)
    assert workbook.sheetnames[:2] == ["汇总结果", "汇总说明"]
    assert workbook["汇总结果"].cell(2, 3).value == "'=1+1"


def test_analysis_preserves_custom_fields_dates_multiple_sheets_and_nonfirst_header(tmp_path):
    path = tmp_path / "虚构资料.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "基础"
    first.append(["姓名", "学号", "政治面貌", "日期"])
    first.append(["测试学生甲", "20260001", "团员", "2026/9/30"])
    second = workbook.create_sheet("补充")
    second.append(["虚构标题"])
    second.append(["姓名", "学号", "竞赛等级"])
    second.append(["测试学生乙", "20260002", "校级"])
    workbook.save(path)

    service = TableAnalysisService()
    all_datasets = service.analyze_all_sheets(path, {"补充": 2})
    assert len(all_datasets) == 2
    base = all_datasets[0]
    assert base.rows[0].values["date"] == "2026-09-30"
    assert any(key.startswith("custom:") for key in base.columns)
    assert all_datasets[1].detected_header == 2


def test_workspace_keeps_current_merge_result_and_fill_ready_dataset():
    result = DatasetMergeService().merge_vertical([
        dataset("来源.xlsx", ["name"], [{"name": "测试学生甲"}]),
    ])
    workspace = DataWorkspaceService()
    workspace.set_merge_result(result)
    assert workspace.current_merge_result is result
    assert workspace.current_dataset is not None
    assert workspace.current_dataset.rows[0].values["name"] == "测试学生甲"
