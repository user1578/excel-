"""考勤统计、来源追溯、Excel 导出和统计页离屏冒烟测试。"""

from __future__ import annotations

from pathlib import Path
import os

import pytest
from openpyxl import load_workbook

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
from PySide6.QtWidgets import QApplication

from app.models.class_record import ClassRecord
from app.models.student import Student
from app.models.task import Task
from app.repositories.database import DatabaseManager
from app.services.excel_export_service import ExcelExportService
from app.services.master_data_service import MasterDataService
from app.services.statistics_service import AttendanceQuery, StatisticsService
from app.services.task_service import TaskService
from app.ui.statistics_page import StatisticsPage, StudentDetailDialog


@pytest.fixture
def statistics_setup(tmp_path):
    database = DatabaseManager(tmp_path / "data" / "database.db")
    database.initialize()
    master = MasterDataService(database)
    master.create_class(ClassRecord("物联网2401", student_count=40))
    master.create_class(ClassRecord("软件2401", student_count=0))
    students = {
        "张三": master.create_student(Student("张三", "20260001", "物联网2401")),
        "李四": master.create_student(Student("李四", "20260002", "物联网2401")),
        "王五": master.create_student(Student("王五", "20260003", "物联网2401")),
        "赵六": master.create_student(Student("赵六", "20260004", "物联网2401")),
        "钱七": master.create_student(Student("钱七", "20260005", "软件2401")),
        "孙八": master.create_student(Student("孙八", "20260006", "软件2401")),
    }
    tasks = TaskService(database)
    september = tasks.create(Task("2026年9月课堂查课"))
    october = tasks.create(Task("2026年10月课堂查课"))
    with database.transaction() as connection:
        source_ids = {}
        for task in (september, october):
            source_ids[task.id] = connection.execute(
                """INSERT INTO source_files (task_id, original_name, original_path, stored_path, file_type, file_size, file_hash, sheet_name, record_count, status)
                VALUES (?, ?, ?, ?, 'xlsx', 1, ?, '考勤', 1, '已导入')""",
                (task.id, f"{task.name}.xlsx", f"{task.name}.xlsx", f"{task.name}.xlsx", f"hash-{task.id}"),
            ).lastrowid
        records = [
            (september.id, "张三", "2026-09-02", "课堂", "迟到", 3),
            (september.id, "张三", "2026-09-03", "课堂", "正常", 1),
            (september.id, "李四", "2026-09-03", "课堂", "缺勤", 2),
            (september.id, "王五", "2026-09-20", "课堂", "请假", 1),
            (september.id, "赵六", "2026-09-20", "课堂", "正常", 1),
            (september.id, "钱七", "2026-09-25", "晚自习", "早退", 2),
            (september.id, "孙八", "2026-10-01", "课堂", "迟到", 1),
            (october.id, "张三", "2026-10-02", "课堂", "缺勤", 1),
            (october.id, "李四", "2026-10-04", "课堂", "正常", 1),
        ]
        for row_number, (task_id, name, record_date, attendance_type, status, count) in enumerate(records, 2):
            student = students[name]
            connection.execute(
                """INSERT INTO attendance_records
                (task_id, source_file_id, source_row_number, date, attendance_type, student_id, student_name, student_number, class_name, status, count, course, remark, raw_data)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Python', '测试备注', '{}')""",
                (task_id, source_ids[task_id], row_number, record_date, attendance_type, student.id, student.name, student.student_number, student.class_name, status, count),
            )
    service = StatisticsService(database)
    return service, ExcelExportService(service, tmp_path / "exports"), master, tasks, students, september, october


def by_name(rows, name):
    return next(row for row in rows if row["name"] == name)


def test_personal_summary_uses_count_and_groups_known_statuses(statistics_setup):
    service, _exporter, _master, _tasks, _students, _september, _october = statistics_setup
    result = service.summarize(AttendanceQuery())
    zhang = by_name(result.personal_rows, "张三")
    assert (zhang["late_count"], zhang["absent_count"], zhang["normal_count"], zhang["abnormal_count"], zhang["record_count"]) == (3, 1, 1, 4, 3)
    qian = by_name(result.personal_rows, "钱七")
    assert (qian["other_count"], qian["abnormal_count"]) == (2, 2)
    assert result.personal_rows[0]["name"] == "张三"


def test_class_summary_uses_configured_or_fallback_student_count(statistics_setup):
    service, _exporter, _master, _tasks, _students, _september, _october = statistics_setup
    rows = service.summarize(AttendanceQuery()).class_rows
    internet = next(row for row in rows if row["class_name"] == "物联网2401")
    software = next(row for row in rows if row["class_name"] == "软件2401")
    assert (internet["class_student_count"], internet["record_student_count"], internet["late_count"], internet["absent_count"], internet["leave_count"], internet["abnormal_count"]) == (40, 4, 3, 3, 1, 7)
    assert (software["class_student_count"], software["record_student_count"], software["other_count"]) == (2, 2, 2)


@pytest.mark.parametrize("query, expected_names", [
    (AttendanceQuery(task_id=1), {"张三", "李四", "王五", "赵六", "钱七", "孙八"}),
    (AttendanceQuery(start_date="2026-09-01", end_date="2026-09-30"), {"张三", "李四", "王五", "赵六", "钱七"}),
    (AttendanceQuery(class_name="软件2401"), {"钱七", "孙八"}),
    (AttendanceQuery(student_id=1), {"张三"}),
    (AttendanceQuery(attendance_type="晚自习"), {"钱七"}),
    (AttendanceQuery(status="其他"), {"钱七"}),
    (AttendanceQuery(task_id=1, start_date="2026-09-01", end_date="2026-09-30", class_name="物联网2401", attendance_type="课堂"), {"张三", "李四", "王五", "赵六"}),
])
def test_filters_can_be_combined(statistics_setup, query, expected_names):
    service, _exporter, _master, _tasks, students, september, _october = statistics_setup
    replacements = {1: september.id}
    if query.student_id == 1:
        query = AttendanceQuery(student_id=students["张三"].id)
    elif query.task_id == 1:
        query = AttendanceQuery(september.id, query.start_date, query.end_date, query.class_name, query.student_id, query.attendance_type, query.status)
    assert {row["name"] for row in service.summarize(query).personal_rows} == expected_names


def test_student_detail_includes_source_traceability_and_empty_result(statistics_setup):
    service, _exporter, _master, _tasks, students, september, _october = statistics_setup
    rows = service.student_detail(AttendanceQuery(task_id=september.id), students["张三"].id)
    assert len(rows) == 2
    assert rows[0]["source_file_name"].endswith(".xlsx")
    assert rows[0]["sheet_name"] == "考勤"
    assert rows[0]["source_row_number"]
    assert service.summarize(AttendanceQuery(start_date="2030-01-01", end_date="2030-01-31")).overview["record_count"] == 0


def test_dates_for_period_are_inclusive_and_validated(statistics_setup):
    service, _exporter, _master, _tasks, _students, _september, _october = statistics_setup
    assert service.dates_for_period("本周", __import__("datetime").date(2026, 9, 2)) == ("2026-08-31", "2026-09-06")
    assert service.dates_for_period("本月", __import__("datetime").date(2026, 9, 2)) == ("2026-09-01", "2026-09-30")
    with pytest.raises(ValueError):
        AttendanceQuery(start_date="2026-10-01", end_date="2026-09-01")


def test_summary_export_is_complete_styled_and_unique(statistics_setup):
    service, exporter, _master, _tasks, _students, september, _october = statistics_setup
    query = AttendanceQuery(task_id=september.id, start_date="2026-09-01", end_date="2026-09-30")
    first = exporter.export_summaries(query)
    second = exporter.export_summaries(query)
    assert first != second and first.exists() and second.exists()
    workbook = load_workbook(first)
    assert workbook.sheetnames == ["个人汇总", "班级汇总", "统计说明"]
    personal = workbook["个人汇总"]
    assert [cell.value for cell in personal[1]][:3] == ["姓名", "学号", "班级"]
    assert personal.freeze_panes == "A2" and personal.auto_filter.ref
    assert workbook["统计说明"]["A1"].value == "统计说明"


def test_student_detail_export_and_no_data_error(statistics_setup):
    service, exporter, _master, _tasks, students, september, _october = statistics_setup
    path = exporter.export_student_detail(AttendanceQuery(task_id=september.id), students["张三"].id, "张三")
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["学生明细", "统计说明"]
    assert [cell.value for cell in workbook["学生明细"][1]][:5] == ["日期", "考勤类型", "课程", "状态", "次数"]
    with pytest.raises(ValueError, match="暂无考勤记录"):
        exporter.export_summaries(AttendanceQuery(start_date="2030-01-01", end_date="2030-01-31"))


def test_statistics_page_smoke_can_query_switch_detail_and_export(statistics_setup, monkeypatch):
    service, exporter, master, tasks, _students, september, _october = statistics_setup
    application = QApplication.instance() or QApplication([])
    page = StatisticsPage(service, exporter, master, tasks)
    page.show(); application.processEvents()
    page.task_box.setCurrentIndex(page.task_box.findData(september.id))
    page.run_query()
    assert page.personal_table.rowCount() == 6
    page.tabs.setCurrentIndex(1)
    assert page.class_table.rowCount() == 2
    monkeypatch.setattr(StudentDetailDialog, "exec", lambda _self: 0)
    page.open_student_detail(0)
    assert page.export_current(show_message=False) is not None
    page.close()
