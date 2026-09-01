"""V2.2.1 模板视觉、大标题模式和 Workbook Fill 自动序号测试。"""

from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PySide6.QtWidgets import QApplication

from app.models.table_dataset import Provenance, TableDataset, TableRow
from app.parsers.workbook_template_analyzer import WorkbookTemplateAnalyzer
from app.services.master_data_service import MasterDataService
from app.services.template_service import TemplateService
from app.services.workbook_fill_service import AUTO_SEQUENCE, SEQUENCE_FILL_BLANK, SEQUENCE_NONE, SEQUENCE_RENUMBER, SKIP_CONFLICTING_ROW, USE_NEW_VALUE, WorkbookFillService
from app.template_engine.schema import FieldSchema, SheetSchema, TemplateSchema
from app.template_engine.styles import WorkbookStyleSchema, business_blue_style, standard_office_style
from app.repositories.database import DatabaseManager
from app.services.data_workspace_service import DataWorkspaceService
from app.ui.dialogs.style_dialog import StyleDialog
from app.ui.workbook_fill_page import WorkbookFillPage


def _master(tmp_path):
    database = DatabaseManager(tmp_path / "v221.db")
    database.initialize()
    return MasterDataService(database)


def _dataset(rows=3):
    return TableDataset(
        ["name"],
        [TableRow({"name": f"测试学生{index}"}, Provenance("虚构.xlsx", "名单", index + 1)) for index in range(1, rows + 1)],
        "虚构.xlsx", "名单", 1, column_labels={"name": "姓名"},
    )


def _fill_template(path, first_sequence=None, first_name=""):
    workbook = Workbook(); sheet = workbook.active; sheet.title = "名单"
    sheet.append(["序号", "姓名", "编号信息"])
    sheet.append([first_sequence, first_name, "业务编号"])
    for cell in sheet[2]:
        cell.font = Font(name="宋体", bold=True, color="112233")
        cell.fill = PatternFill("solid", fgColor="DDEEFF")
        cell.border = Border(left=Side("thin", color="123456"), right=Side("thin", color="123456"), top=Side("thin", color="123456"), bottom=Side("thin", color="123456"))
        cell.alignment = Alignment(horizontal="center")
    sheet.row_dimensions[2].height = 25
    sheet.column_dimensions["A"].width = 13
    workbook.save(path)


def test_standard_office_defaults_and_semantic_widths(tmp_path):
    style = standard_office_style()
    assert (style.show_gridlines, style.auto_filter, style.body_horizontal_alignment, style.body_vertical_alignment) == (True, False, "center", "center")
    assert (business_blue_style().auto_filter, business_blue_style().show_gridlines) == (True, False)
    schema = TemplateSchema("学生名单", default_rows=1, sheets=[SheetSchema("录入", [
        FieldSchema("序号"), FieldSchema("姓名"), FieldSchema("学号"), FieldSchema("家庭住址"), FieldSchema("备注", column_width=30),
    ])])
    artifact = TemplateService(_master(tmp_path), tmp_path / "templates").create(schema)
    sheet = load_workbook(artifact.workbook_path)["录入"]
    assert (sheet.sheet_view.showGridLines, sheet.auto_filter.ref, sheet["A2"].alignment.horizontal) == (True, None, "center")
    assert (sheet.column_dimensions["A"].width, sheet.column_dimensions["B"].width, sheet.column_dimensions["C"].width, sheet.column_dimensions["D"].width, sheet.column_dimensions["E"].width) == (8, 12, 18, 32, 30)


def test_title_modes_keep_real_header_rows_and_legacy_schema_compatibility(tmp_path):
    master = _master(tmp_path); service = TemplateService(master, tmp_path / "templates")
    base = [SheetSchema("录入", [FieldSchema("姓名"), FieldSchema("学号")])]
    title = TemplateSchema("物联网工程班学生名单", default_rows=1, sheets=base, style=WorkbookStyleSchema(title_mode="template_name", show_main_title=True))
    custom = TemplateSchema("报名表", default_rows=1, sheets=base, style=WorkbookStyleSchema(title_mode="custom", show_main_title=True, main_title="自定义登记表", auto_filter=True))
    plain = TemplateSchema("简单名单", default_rows=1, sheets=base, style=WorkbookStyleSchema(title_mode="none", show_main_title=False))
    title_sheet = load_workbook(service.create(title).workbook_path)["录入"]
    custom_sheet = load_workbook(service.create(custom).workbook_path)["录入"]
    plain_sheet = load_workbook(service.create(plain).workbook_path)["录入"]
    legacy = WorkbookStyleSchema.from_dict({"show_main_title": True, "main_title": "旧标题"})
    assert (title_sheet["A1"].value, title_sheet["A2"].value, title_sheet.freeze_panes) == ("物联网工程班学生名单", "姓名", None)
    assert custom_sheet.auto_filter.ref == "A2:B3" and custom_sheet["A1"].value == "自定义登记表"
    assert (plain_sheet["A1"].value, plain_sheet.freeze_panes, legacy.title_mode) == ("姓名", None, "custom")


def test_sequence_detection_modes_and_preserves_template_style(tmp_path):
    template = tmp_path / "学校名单.xlsx"; _fill_template(template, 100)
    analysis = WorkbookTemplateAnalyzer().analyze(template, "名单", 1)
    service = WorkbookFillService(tmp_path / "exports")
    mappings = service.default_mappings(analysis, _dataset())
    assert mappings["序号"] == AUTO_SEQUENCE and "编号信息" not in mappings
    result = service.fill(analysis, _dataset(), mappings, USE_NEW_VALUE, sequence_start=1, sequence_mode=SEQUENCE_FILL_BLANK)
    sheet = load_workbook(result.output_path)["名单"]
    assert [sheet.cell(row, 1).value for row in (2, 3, 4)] == [100, 2, 3]
    assert (sheet["A3"].font.bold, sheet["A3"].fill.fgColor.rgb, sheet["A3"].border.left.style, sheet.row_dimensions[3].height, sheet.column_dimensions["A"].width) == (True, "00DDEEFF", "thin", 25, 13)
    renumber = service.fill(analysis, _dataset(), mappings, USE_NEW_VALUE, sequence_start=1, sequence_mode=SEQUENCE_RENUMBER)
    untouched = service.fill(analysis, _dataset(), mappings, USE_NEW_VALUE, sequence_start=1, sequence_mode=SEQUENCE_NONE)
    assert [load_workbook(renumber.output_path)["名单"].cell(row, 1).value for row in (2, 3, 4)] == [1, 2, 3]
    assert load_workbook(untouched.output_path)["名单"]["A2"].value == 100


def test_sequence_stays_continuous_after_skipped_row_and_can_be_mapped_manually(tmp_path):
    template = tmp_path / "排名模板.xlsx"; _fill_template(template, None, "原姓名")
    workbook = load_workbook(template); workbook["名单"]["A1"] = "排名编号"; workbook.save(template)
    analysis = WorkbookTemplateAnalyzer().analyze(template, "名单", 1)
    service = WorkbookFillService(tmp_path / "exports")
    mappings = {"排名编号": AUTO_SEQUENCE, "姓名": "name"}
    preview = service.preview(analysis, _dataset(), mappings, sequence_start=10, sequence_mode=SEQUENCE_RENUMBER)
    result = service.fill(analysis, _dataset(), mappings, SKIP_CONFLICTING_ROW, sequence_start=10, sequence_mode=SEQUENCE_RENUMBER)
    sheet = load_workbook(result.output_path)["名单"]
    assert (preview.sequence_target, preview.sequence_start, result.skipped_rows) == ("排名编号", 10, 1)
    assert (sheet["A3"].value, sheet["A4"].value) == (10, 11)


def test_sequence_controls_and_title_mode_are_available_offscreen(tmp_path):
    application = QApplication.instance() or QApplication([])
    template = tmp_path / "界面模板.xlsx"; _fill_template(template)
    page = WorkbookFillPage(DataWorkspaceService(), _master(tmp_path))
    page.analysis = WorkbookTemplateAnalyzer().analyze(template, "名单", 1)
    page.dataset = _dataset()
    page._refresh_mapping_table(); application.processEvents()
    assert page.sequence_enabled.isChecked() and page.sequence_start.value() == 1
    page.sequence_enabled.setChecked(False)
    assert AUTO_SEQUENCE not in page._mappings().values()
    dialog = StyleDialog(standard_office_style())
    dialog.title_mode.setCurrentIndex(dialog.title_mode.findData("custom")); dialog.title.setText("自定义标题")
    style = dialog.result_style()
    assert (style.title_mode, style.show_main_title, style.main_title) == ("custom", True, "自定义标题")
    dialog.close(); page.close()
