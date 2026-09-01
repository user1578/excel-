"""V2.2.2 冻结窗格与新生成表格尺寸的回归测试。"""

from __future__ import annotations

import os
from dataclasses import replace

from openpyxl import load_workbook
from PySide6.QtWidgets import QApplication

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from app.models.class_record import ClassRecord
from app.models.student import Student
from app.repositories.database import DatabaseManager
from app.services.class_export_service import ClassExportService, ExportColumn, SOURCE_CORE, SOURCE_EXTRA
from app.services.master_data_service import MasterDataService
from app.services.template_service import TemplateService
from app.template_engine.schema import FieldSchema, SheetSchema, TemplateSchema
from app.template_engine.styles import WorkbookStyleSchema, business_blue_style, standard_office_style
from app.ui.dialogs.style_dialog import StyleDialog
from app.ui.template_page import TemplatePage


def _master(tmp_path):
    database = DatabaseManager(tmp_path / "v222.db")
    database.initialize()
    return MasterDataService(database)


def _template(style, fields, rows=2):
    return TemplateSchema("尺寸测试", default_rows=rows, sheets=[SheetSchema("录入", fields)], style=style)


def test_freeze_modes_keep_title_position_and_support_legacy_style(tmp_path):
    service = TemplateService(_master(tmp_path), tmp_path / "templates")
    fields = [FieldSchema("姓名"), FieldSchema("学号")]
    plain = load_workbook(service.create(_template(standard_office_style(), fields)).workbook_path)["录入"]
    frozen = load_workbook(service.create(_template(replace(standard_office_style(), freeze_mode="header"), fields)).workbook_path)["录入"]
    titled = load_workbook(service.create(_template(WorkbookStyleSchema(title_mode="template_name", show_main_title=True, freeze_mode="header"), fields)).workbook_path)["录入"]
    legacy = WorkbookStyleSchema.from_dict({"freeze_header": True})

    assert (plain.freeze_panes, frozen.freeze_panes, titled.freeze_panes) == (None, "A2", "A3")
    assert legacy.freeze_mode == "header" and legacy.freeze_header
    assert business_blue_style().freeze_mode == "header"


def test_freeze_choice_only_changes_this_generation_style():
    original = replace(standard_office_style(), freeze_mode="ask")
    frozen = TemplatePage._style_with_freeze_choice(original, True)
    unfrozen = TemplatePage._style_with_freeze_choice(original, False)
    assert (original.freeze_mode, frozen.freeze_mode, frozen.freeze_header) == ("ask", "header", True)
    assert (unfrozen.freeze_mode, unfrozen.freeze_header) == ("none", False)


def test_auto_fit_counts_chinese_width_applies_priority_and_wraps_header(tmp_path):
    service = TemplateService(_master(tmp_path), tmp_path / "templates")
    long_header = "申请生源地信用助学贷款县市区详细行政区划名称"
    schema = _template(standard_office_style(), [
        FieldSchema("姓名"), FieldSchema("学号"), FieldSchema(long_header), FieldSchema("备注", column_width=31),
    ])
    sheet = load_workbook(service.create(schema).workbook_path)["录入"]

    assert (sheet.column_dimensions["A"].width, sheet.column_dimensions["B"].width, sheet.column_dimensions["D"].width) == (12, 18, 31)
    assert sheet.column_dimensions["C"].width == 35
    assert sheet.row_dimensions[1].height > standard_office_style().header_row_height


def test_blank_template_and_disabled_auto_fit_keep_base_body_height(tmp_path):
    service = TemplateService(_master(tmp_path), tmp_path / "templates")
    normal = load_workbook(service.create(_template(standard_office_style(), [FieldSchema("姓名"), FieldSchema("备注")], rows=3)).workbook_path)["录入"]
    fixed_style = replace(standard_office_style(), auto_fit=False, default_column_width=14)
    fixed = load_workbook(service.create(_template(fixed_style, [FieldSchema("未使用的长字段名称", default_value="这是一段关闭自动调整后也不应增加行高的长文本")], rows=2)).workbook_path)["录入"]

    assert [normal.row_dimensions[row].height for row in (2, 3, 4)] == [22, 22, 22]
    assert (fixed.column_dimensions["A"].width, fixed.row_dimensions[2].height, fixed.row_dimensions[3].height) == (14, 22, 22)


def test_class_export_measures_actual_extra_values_and_expands_wrapped_body(tmp_path):
    master = _master(tmp_path)
    master.create_class(ClassRecord("测试班2401"))
    student = master.create_student(Student("测试学生", "20260001", "测试班2401"))
    master.set_student_extra_field(student.id, "家庭住址", "广东省广州市海珠区新港中路测试学生家庭详细住址信息")
    output = ClassExportService(master, tmp_path / "exports").export(
        "测试班2401", [student], [
            ExportColumn("姓名", SOURCE_CORE, "name"),
            ExportColumn("家庭住址", SOURCE_EXTRA, "家庭住址"),
        ], "", standard_office_style(),
    )
    sheet = load_workbook(output)["学生名单"]

    assert sheet.freeze_panes is None
    assert 12 < sheet.column_dimensions["B"].width <= 35
    assert sheet.row_dimensions[2].height > standard_office_style().body_row_height


def test_style_dialog_exposes_freeze_modes_and_auto_fit_offscreen():
    application = QApplication.instance() or QApplication([])
    dialog = StyleDialog(standard_office_style())
    values = [dialog.freeze_mode.itemData(index) for index in range(dialog.freeze_mode.count())]
    dialog.freeze_mode.setCurrentIndex(dialog.freeze_mode.findData("header"))
    dialog.auto_fit.setChecked(False)
    result = dialog.result_style()

    assert values == ["none", "header", "ask"]
    assert (result.freeze_mode, result.freeze_header, result.auto_fit) == ("header", True, False)
    dialog.close()
    application.processEvents()
