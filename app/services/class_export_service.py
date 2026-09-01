"""按班级学生资料生成可配置的普通名单式 Excel。"""

from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.table_dataset import Provenance, TableDataset, TableRow
from app.models.student import Student
from app.repositories.table_export_scheme_repository import TableExportScheme, TableExportSchemeRepository
from app.services.master_data_service import MasterDataService
from app.utils.excel_safety import safe_excel_value


CORE_LABELS = {"name": "姓名", "student_number": "学号", "class_name": "班级", "major": "专业", "grade": "年级", "phone": "联系电话", "dormitory": "寝室", "remark": "备注"}
SOURCE_AUTO = "auto_number"
SOURCE_CORE = "core"
SOURCE_EXTRA = "extra"
SOURCE_BLANK = "blank"
SOURCE_FIXED = "fixed"


@dataclass(frozen=True)
class ExportColumn:
    title: str
    source_type: str
    source_field: str = ""
    fixed_value: str = ""


class ClassExportService:
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))

    def __init__(self, master: MasterDataService, exports_directory: str | Path | None = None) -> None:
        self.master = master
        self.exports_directory = Path(exports_directory or Path(__file__).resolve().parents[2] / "exports")
        self.schemes = TableExportSchemeRepository(master.students.database)

    def students_dataset(self, students: Iterable[Student], source_name: str = "班级学生") -> TableDataset:
        selected = list(students)
        extra_names: dict[str, str] = {}
        extra_values: dict[int, dict[str, dict[str, str]]] = {}
        for student in selected:
            if student.id is None:
                continue
            values = self.master.get_student_extra_fields(student.id)
            extra_values[student.id] = values
            extra_names.update({key: item["name"] for key, item in values.items()})
        columns = list(CORE_LABELS) + [f"extra:{key}" for key in sorted(extra_names)]
        labels = dict(CORE_LABELS) | {f"extra:{key}": name for key, name in extra_names.items()}
        rows: list[TableRow] = []
        for index, student in enumerate(selected, 1):
            values = {field: getattr(student, field) or "" for field in CORE_LABELS}
            values.update({f"extra:{key}": extra_values.get(student.id or -1, {}).get(key, {}).get("value", "") for key in extra_names})
            rows.append(TableRow(values, Provenance(source_name, None, index + 1)))
        return TableDataset(columns, rows, source_name, None, 1, column_labels=labels, custom_fields={key for key in columns if key.startswith("extra:")})

    def export(self, class_name: str, students: Iterable[Student], columns: list[ExportColumn], title: str = "") -> Path:
        if not columns:
            raise ValueError("请至少保留一个导出列。")
        selected = list(students)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "学生名单"
        sheet.sheet_view.showGridLines = False
        header_row = 2 if title.strip() else 1
        if title.strip():
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
            cell = sheet.cell(1, 1, safe_excel_value(title.strip()))
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[1].height = 28
        for column_index, definition in enumerate(columns, 1):
            cell = sheet.cell(header_row, column_index, safe_excel_value(definition.title))
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_index, student in enumerate(selected, 1):
            extras = self.master.get_student_extra_fields(student.id) if student.id else {}
            for column_index, definition in enumerate(columns, 1):
                value = self._column_value(definition, student, extras, row_index)
                cell = sheet.cell(header_row + row_index, column_index, safe_excel_value(value))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = self.BORDER
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{max(header_row, header_row + len(selected))}"
        for index, definition in enumerate(columns, 1):
            width = max(len(definition.title), *(len(str(self._column_value(definition, student, self.master.get_student_extra_fields(student.id) if student.id else {}, row + 1))) for row, student in enumerate(selected)), 8) + 2
            sheet.column_dimensions[get_column_letter(index)].width = min(width, 30)
        self.exports_directory.mkdir(parents=True, exist_ok=True)
        output = self._unique_path(f"{self._safe_name(class_name)}_{self._safe_name(title) if title.strip() else '学生名单'}")
        workbook.save(output)
        load_workbook(output).close()
        return output

    def save_scheme(self, name: str, title: str, columns: list[ExportColumn]) -> TableExportScheme:
        if self.schemes.get_by_name(name):
            raise ValueError("方案名称已存在；请使用其他名称，系统不会静默覆盖。")
        return self.schemes.create(TableExportScheme(name.strip(), title, [asdict(item) for item in columns]))

    def list_schemes(self) -> list[TableExportScheme]:
        return self.schemes.list_all()

    def delete_scheme(self, scheme_id: int) -> bool:
        return self.schemes.delete(scheme_id)

    @staticmethod
    def _column_value(definition: ExportColumn, student: Student, extras: dict[str, dict[str, str]], sequence: int):
        if definition.source_type == SOURCE_AUTO:
            return sequence
        if definition.source_type == SOURCE_CORE:
            return getattr(student, definition.source_field, "") or ""
        if definition.source_type == SOURCE_EXTRA:
            return extras.get(definition.source_field, {}).get("value", "")
        if definition.source_type == SOURCE_FIXED:
            return definition.fixed_value
        return ""

    def _unique_path(self, stem: str) -> Path:
        base = f"{stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        path = self.exports_directory / f"{base}.xlsx"
        suffix = 2
        while path.exists():
            path = self.exports_directory / f"{base}_{suffix}.xlsx"
            suffix += 1
        return path

    @staticmethod
    def _safe_name(value: str) -> str:
        return "".join("_" if char in '\\/:*?\"<>|' else char for char in value).strip() or "未命名"
