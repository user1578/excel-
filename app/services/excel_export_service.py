"""将当前统计条件的完整数据库结果导出为 xlsx。"""

from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.statistics_service import AttendanceQuery, StatisticsService
from app.utils.excel_safety import safe_excel_value


class ExcelExportService:
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    THIN_BORDER = Border(bottom=Side(style="thin", color="D9E2F3"))

    def __init__(self, statistics: StatisticsService, exports_directory: str | Path | None = None) -> None:
        self.statistics = statistics
        self.exports_directory = Path(exports_directory or Path(__file__).resolve().parents[2] / "exports")

    def export_summaries(self, query: AttendanceQuery) -> Path:
        result = self.statistics.summarize(query)
        if not result.overview["record_count"]:
            raise ValueError("当前条件下暂无考勤记录，无法导出。")
        workbook = Workbook()
        workbook.remove(workbook.active)
        self._write_table(workbook.create_sheet("个人汇总"), "个人汇总", [
            ("姓名", "name"), ("学号", "student_number"), ("班级", "class_name"), ("迟到次数", "late_count"),
            ("缺勤次数", "absent_count"), ("请假次数", "leave_count"), ("正常次数", "normal_count"),
            ("其他次数", "other_count"), ("异常总次数", "abnormal_count"), ("记录总数", "record_count"),
        ], result.personal_rows)
        self._write_table(workbook.create_sheet("班级汇总"), "班级汇总", [
            ("班级", "class_name"), ("班级人数", "class_student_count"), ("记录学生数", "record_student_count"),
            ("迟到次数", "late_count"), ("缺勤次数", "absent_count"), ("请假次数", "leave_count"),
            ("正常次数", "normal_count"), ("其他次数", "other_count"), ("异常总次数", "abnormal_count"), ("记录总数", "record_count"),
        ], result.class_rows)
        self._write_metadata(workbook.create_sheet("统计说明"), query)
        return self._save(workbook, f"个人班级汇总_{self._period_label(query)}")

    def export_student_detail(self, query: AttendanceQuery, student_id: int, student_name: str) -> Path:
        rows = self.statistics.student_detail(query, student_id)
        if not rows:
            raise ValueError("当前条件下该学生暂无考勤记录，无法导出。")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "学生明细"
        self._write_table(sheet, "学生明细", [
            ("日期", "date"), ("考勤类型", "attendance_type"), ("课程", "course"), ("状态", "status"), ("次数", "count"),
            ("任务", "task_name"), ("来源文件", "source_file_name"), ("工作表", "sheet_name"), ("原始行号", "source_row_number"), ("备注", "remark"),
        ], rows)
        self._write_metadata(workbook.create_sheet("统计说明"), query)
        return self._save(workbook, f"学生明细_{self._safe_name(student_name)}_{self._period_label(query)}")

    def _write_table(self, sheet, title: str, columns: list[tuple[str, str]], rows: list[dict[str, Any]]) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.append([label for label, _key in columns])
        for cell in sheet[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            sheet.append([safe_excel_value(row.get(key, "") if row.get(key) is not None else "") for _label, key in columns])
        for cells in sheet.iter_rows(min_row=2):
            for cell in cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = self.THIN_BORDER
                if isinstance(cell.value, int):
                    cell.number_format = "#,##0"
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(rows) + 1)}"
        for index, (label, _key) in enumerate(columns, 1):
            content_length = max([len(label), *[len(str(row.get(_key, "") or "")) for row in rows]], default=len(label))
            sheet.column_dimensions[get_column_letter(index)].width = min(max(content_length + 2, 10), 30)
        sheet.row_dimensions[1].height = 24

    def _write_metadata(self, sheet, query: AttendanceQuery) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.append(["统计说明", "筛选值"])
        names = {"task_id": "任务 ID", "start_date": "开始日期", "end_date": "结束日期", "class_name": "班级", "student_id": "学生 ID", "attendance_type": "考勤类型", "status": "考勤状态"}
        sheet.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        for key, value in asdict(query).items():
            sheet.append([names[key], safe_excel_value(value if value not in (None, "") else "全部")])
        for cell in sheet[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 32

    def _save(self, workbook: Workbook, stem: str) -> Path:
        self.exports_directory.mkdir(parents=True, exist_ok=True)
        path = self.exports_directory / f"{stem}.xlsx"
        if path.exists():
            path = self.exports_directory / f"{stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        suffix = 2
        while path.exists():
            path = self.exports_directory / f"{stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{suffix}.xlsx"
            suffix += 1
        workbook.save(path)
        return path

    @staticmethod
    def _period_label(query: AttendanceQuery) -> str:
        if query.start_date and query.end_date and query.start_date[:7] == query.end_date[:7]:
            return query.start_date[:7].replace("-", "")
        return "全部" if not query.start_date and not query.end_date else "自定义"

    @staticmethod
    def _safe_name(value: str) -> str:
        return "".join("_" if character in '\\/:*?\"<>|' else character for character in value).strip() or "未命名"
