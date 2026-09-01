"""资料汇总结果的安全 xlsx 导出。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.merge_models import MergeResult
from app.utils.excel_safety import safe_excel_value


class MergeExportService:
    def __init__(self, exports_directory: str | Path | None = None) -> None:
        self.exports_directory = Path(exports_directory or Path(__file__).resolve().parents[2] / "exports")

    def export(self, result: MergeResult, allow_unresolved: bool = False) -> Path:
        if result.unresolved_conflicts and not allow_unresolved:
            raise ValueError("仍有未解决字段冲突；请先解决，或在界面中明确确认后导出。")
        workbook = Workbook()
        summary = workbook.active
        summary.title = "汇总结果"
        self._write_result(summary, result)
        self._write_notes(workbook.create_sheet("汇总说明"), result)
        if result.conflicts:
            self._write_conflicts(workbook.create_sheet("字段冲突"), result)
        self.exports_directory.mkdir(parents=True, exist_ok=True)
        path = self._unique_path()
        workbook.save(path)
        load_workbook(path).close()
        return path

    @staticmethod
    def _write_result(sheet, result: MergeResult) -> None:
        sheet.append([result.column_labels[key] for key in result.columns])
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for record in result.records:
            sheet.append([safe_excel_value(record.values.get(key, "")) for key in result.columns])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(max(1, len(result.columns)))}{max(1, len(result.records) + 1)}"
        for column, key in enumerate(result.columns, 1):
            label = result.column_labels[key]
            width = max(len(label), *(len(str(record.values.get(key, "") or "")) for record in result.records)) + 2
            sheet.column_dimensions[get_column_letter(column)].width = min(max(width, 10), 30)

    @staticmethod
    def _write_notes(sheet, result: MergeResult) -> None:
        notes: list[tuple[str, Any]] = [
            ("汇总时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("合并模式", "直接纵向合并" if result.mode.value == "vertical" else "按学生关联合并"),
            ("总行数", len(result.records)),
            ("成功关联数量", len(result.records) - MergeExportService._unlinked_count(result)),
            ("未关联数量", MergeExportService._unlinked_count(result)),
            ("冲突数量", len(result.conflicts)),
            ("已解决冲突数量", len(result.resolved_conflicts)),
            ("未解决冲突数量", len(result.unresolved_conflicts)),
        ]
        for source_file, source_sheet in result.source_datasets:
            notes.append(("来源文件 / Sheet", f"{source_file} / {source_sheet or 'CSV'}"))
        for key, value in notes:
            sheet.append([key, safe_excel_value(value)])
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 42

    @staticmethod
    def _unlinked_count(result: MergeResult) -> int:
        """unresolved 与 unmatched 都代表未成功关联，索引去重避免重复统计。"""
        unresolved = set(result.unresolved_record_indexes)
        unmatched = {
            index for index, record in enumerate(result.records)
            if getattr(record, "match_status", None) == "unmatched"
        }
        return len(unresolved | unmatched)

    @staticmethod
    def _write_conflicts(sheet, result: MergeResult) -> None:
        sheet.append(["编号", "字段", "值 A", "来源 A", "值 B", "来源 B", "解决状态", "最终值"])
        for conflict in result.conflicts:
            source_a = f"{conflict.source_a.source_file}/{conflict.source_a.source_sheet or 'CSV'}:{conflict.source_a.source_row}"
            source_b = f"{conflict.source_b.source_file}/{conflict.source_b.source_sheet or 'CSV'}:{conflict.source_b.source_row}"
            sheet.append([conflict.id, conflict.field, safe_excel_value(conflict.value_a), source_a, safe_excel_value(conflict.value_b), source_b, conflict.resolution.value, safe_excel_value(conflict.resolved_value or "")])
        sheet.freeze_panes = "A2"

    def _unique_path(self) -> Path:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = self.exports_directory / f"资料汇总_{stamp}.xlsx"
        suffix = 2
        while path.exists():
            path = self.exports_directory / f"资料汇总_{stamp}_{suffix}.xlsx"
            suffix += 1
        return path
