"""将 TableDataset 另存填入现有 .xlsx 模板，绝不原地修改模板。"""

from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

from app.models.fill_models import FillPreview, FillResult, TemplateAnalysis
from app.models.table_dataset import TableDataset
from app.parsers.field_detector import ALIASES
from app.utils.excel_safety import safe_excel_value
from app.utils.value_normalizer import normalize_column_name


KEEP_EXISTING = "keep"
USE_NEW_VALUE = "overwrite"
SKIP_CONFLICTING_ROW = "skip_row"


class MergedCellWriteError(ValueError):
    """用户映射指向合并区域的非左上角单元格。"""


class WorkbookFillService:
    def __init__(self, exports_directory: str | Path | None = None) -> None:
        self.exports_directory = Path(exports_directory or Path(__file__).resolve().parents[2] / "exports")

    def default_mappings(self, analysis: TemplateAnalysis, dataset: TableDataset) -> dict[str, str]:
        result: dict[str, str] = {}
        source_by_label = {normalize_column_name(dataset.display_label(key)): key for key in dataset.columns}
        for target, _column in analysis.target_columns.items():
            normalized = normalize_column_name(target)
            if normalized in source_by_label:
                result[target] = source_by_label[normalized]
                continue
            for field, aliases in ALIASES.items():
                if target in aliases and field.value in dataset.columns:
                    result[target] = field.value
                    break
        return result

    def preview(self, analysis: TemplateAnalysis, dataset: TableDataset, mappings: dict[str, str]) -> FillPreview:
        self._validate_mappings(analysis, dataset, mappings)
        workbook = load_workbook(analysis.template_path, data_only=False)
        try:
            sheet = workbook[analysis.sheet_name]
            warnings = self._merged_warnings(sheet, analysis, len(dataset.rows), mappings)
            conflicts = 0
            data_start = analysis.header_row + 1
            for index, _row in enumerate(dataset.rows):
                excel_row = data_start + index
                for target in mappings:
                    cell = sheet.cell(excel_row, analysis.target_columns[target])
                    if cell.value not in (None, ""):
                        conflicts += 1
            return FillPreview(len(dataset.rows), dict(mappings), conflicts, warnings)
        finally:
            workbook.close()

    def fill(
        self,
        analysis: TemplateAnalysis,
        dataset: TableDataset,
        mappings: dict[str, str],
        existing_value_strategy: str = KEEP_EXISTING,
    ) -> FillResult:
        if existing_value_strategy not in {KEEP_EXISTING, USE_NEW_VALUE, SKIP_CONFLICTING_ROW}:
            raise ValueError("未知的模板已有值处理策略。")
        self._validate_mappings(analysis, dataset, mappings)
        workbook = load_workbook(analysis.template_path, data_only=False)
        try:
            sheet = workbook[analysis.sheet_name]
            warnings = self._merged_warnings(sheet, analysis, len(dataset.rows), mappings)
            if warnings:
                raise MergedCellWriteError("；".join(warnings))
            data_start = analysis.header_row + 1
            written_rows = skipped_rows = preserved_cells = 0
            for offset, source_row in enumerate(dataset.rows):
                row_number = data_start + offset
                if row_number > data_start:
                    self._copy_template_row(sheet, data_start, row_number)
                targets = [(target, analysis.target_columns[target], source_row.values.get(source_key, "")) for target, source_key in mappings.items()]
                existing = [sheet.cell(row_number, column).value not in (None, "") for _target, column, _value in targets]
                if existing_value_strategy == SKIP_CONFLICTING_ROW and any(existing):
                    skipped_rows += 1
                    continue
                wrote = False
                for (_target, column, value), has_existing in zip(targets, existing):
                    if has_existing and existing_value_strategy == KEEP_EXISTING:
                        preserved_cells += 1
                        continue
                    sheet.cell(row_number, column).value = safe_excel_value(value)
                    wrote = True
                if wrote:
                    written_rows += 1
            self.exports_directory.mkdir(parents=True, exist_ok=True)
            output = self._unique_output_path(analysis.template_path)
            workbook.save(output)
        finally:
            workbook.close()
        load_workbook(output, data_only=False).close()
        return FillResult(output, written_rows, skipped_rows, preserved_cells)

    @staticmethod
    def _validate_mappings(analysis: TemplateAnalysis, dataset: TableDataset, mappings: dict[str, str]) -> None:
        if not mappings:
            raise ValueError("请至少配置一个字段映射。")
        unknown_targets = set(mappings) - set(analysis.target_columns)
        unknown_sources = set(mappings.values()) - set(dataset.columns)
        if unknown_targets or unknown_sources:
            raise ValueError("字段映射包含不存在的模板字段或数据源字段。")

    @staticmethod
    def _copy_template_row(sheet, source_row: int, target_row: int) -> None:
        source_height = sheet.row_dimensions[source_row].height
        if source_height is not None:
            sheet.row_dimensions[target_row].height = source_height
        for column in range(1, sheet.max_column + 1):
            source = sheet.cell(source_row, column)
            target = sheet.cell(target_row, column)
            if source.has_style:
                target._style = copy(source._style)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
            if source.data_type == "f":
                target.value = Translator(source.value, origin=source.coordinate).translate_formula(target.coordinate)

    @staticmethod
    def _merged_warnings(sheet, analysis: TemplateAnalysis, rows: int, mappings: dict[str, str]) -> list[str]:
        warnings: list[str] = []
        for offset in range(rows):
            row = analysis.header_row + 1 + offset
            for target in mappings:
                column = analysis.target_columns[target]
                for merged in sheet.merged_cells.ranges:
                    if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
                        if (row, column) != (merged.min_row, merged.min_col):
                            warnings.append(f"字段“{target}”位于合并单元格 {merged.coord} 的非左上角，不能写入")
        return warnings

    def _unique_output_path(self, template_path: Path) -> Path:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = f"{template_path.stem}_已填写_{stamp}"
        path = self.exports_directory / f"{base}.xlsx"
        suffix = 2
        while path.exists():
            path = self.exports_directory / f"{base}_{suffix}.xlsx"
            suffix += 1
        return path
