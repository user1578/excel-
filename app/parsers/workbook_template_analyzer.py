"""只读分析 .xlsx 模板的工作表与目标表头。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from app.models.fill_models import TemplateAnalysis
from app.parsers.excel_reader import read_raw
from app.parsers.header_detector import detect_header
from app.utils.value_normalizer import normalize_text


class WorkbookTemplateAnalyzer:
    def sheets(self, template_path: str | Path) -> list[str]:
        path = self._validate(template_path)
        workbook = load_workbook(path, read_only=True)
        try:
            return workbook.sheetnames
        finally:
            workbook.close()

    def analyze(self, template_path: str | Path, sheet_name: str, header_row: int | None = None) -> TemplateAnalysis:
        path = self._validate(template_path)
        frame = read_raw(path, sheet_name)
        row_number = header_row or detect_header(frame)[0] + 1
        workbook = load_workbook(path, data_only=False)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError("所选工作表不存在。")
            sheet = workbook[sheet_name]
            if row_number < 1 or row_number > sheet.max_row:
                raise ValueError("表头行不在模板范围内。")
            target_columns: dict[str, int] = {}
            non_anchors: set[int] = set()
            for column in range(1, sheet.max_column + 1):
                cell = sheet.cell(row_number, column)
                if isinstance(cell, MergedCell):
                    non_anchors.add(column)
                    continue
                label = normalize_text(cell.value)
                if label:
                    unique = label if label not in target_columns else f"{label}（第{column}列）"
                    target_columns[unique] = column
        finally:
            workbook.close()
        if not target_columns:
            raise ValueError("模板表头行没有可映射字段。")
        return TemplateAnalysis(path, sheet_name, row_number, target_columns, non_anchors)

    @staticmethod
    def _validate(template_path: str | Path) -> Path:
        path = Path(template_path)
        if path.suffix.lower() != ".xlsx":
            raise ValueError("表格填充当前仅支持 .xlsx 模板。")
        if not path.exists():
            raise ValueError("模板文件不存在。")
        return path
