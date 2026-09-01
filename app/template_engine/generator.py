"""本地 TemplateSchema 到 xlsx 的安全生成器。"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from app.template_engine.schema import FIELD_REFERENCE, FieldSchema, TemplateSchema
from app.template_engine.styles import ExcelStyleRenderer


class TemplateGenerator:

    def generate(self, schema: TemplateSchema, output_path: Path, classes: list[str] | None = None, dormitories: list[str] | None = None) -> Path:
        schema.validate()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        renderer = ExcelStyleRenderer(schema.style)
        system_data: dict[str, list[str]] = {}
        for sheet_schema in schema.sheets:
            sheet = workbook.create_sheet(sheet_schema.name)
            self._write_input_sheet(sheet, sheet_schema.fields, schema.default_rows, system_data, classes or [], dormitories or [], renderer, schema.template_name)
        if system_data:
            system_sheet = workbook.create_sheet("_系统数据")
            self._write_system_data(system_sheet, system_data)
            system_sheet.sheet_state = "hidden"
        self._write_instructions(workbook.create_sheet("使用说明"), schema, renderer)
        workbook.save(output_path)
        return output_path

    def _write_input_sheet(self, sheet, fields: list[FieldSchema], rows: int, system_data: dict[str, list[str]], classes: list[str], dormitories: list[str], renderer: ExcelStyleRenderer, template_name: str) -> None:
        renderer.prepare_sheet(sheet)
        header_row = renderer.write_title(sheet, len(fields), template_name)
        data_start = header_row + 1
        field_columns = {field.name: get_column_letter(index) for index, field in enumerate(fields, 1)}
        labels = [field.name for field in fields]
        renderer.style_header(sheet, header_row, labels)
        for index, field in enumerate(fields, 1):
            cell = sheet.cell(header_row, index)
            if field.required:
                cell.comment = __import__("openpyxl").comments.Comment("必填字段", "Excel资料整理助手")
        for row in range(data_start, data_start + rows):
            for index, field in enumerate(fields, 1):
                cell = sheet.cell(row, index)
                renderer.style_body_cell(cell)
                if field.field_type == "formula":
                    cell.value = self._formula(field.formula or "", field_columns, row)
                elif field.default_value is not None:
                    cell.value = field.default_value
                if field.field_type == "date":
                    cell.number_format = "yyyy-mm-dd"
                elif field.field_type == "percentage":
                    cell.number_format = "0.00%"
                elif field.field_type == "formula" and "率" in field.name:
                    cell.number_format = "0.00%"
                elif field.field_type == "integer":
                    cell.number_format = "0"
                elif field.field_type == "decimal":
                    cell.number_format = "0.00"
        for index, field in enumerate(fields, 1):
            if field.field_type == "select" or field.data_source in {"classes", "dormitories"}:
                options = field.options or (classes if field.data_source == "classes" else dormitories)
                key = f"{sheet.title}_{index}_{field.name}"
                system_data[key] = options or [""]
                source_column = get_column_letter(len(system_data))
                formula = f"'_系统数据'!${source_column}$1:${source_column}${len(system_data[key])}"
                validation = DataValidation(type="list", formula1=formula, allow_blank=field.allow_blank)
                sheet.add_data_validation(validation)
                validation.add(f"{get_column_letter(index)}{data_start}:{get_column_letter(index)}{data_start + rows - 1}")
            if field.required:
                renderer.mark_required(sheet.cell(header_row, index), (sheet.cell(row, index) for row in range(data_start, data_start + rows)))
        renderer.size_table(sheet, header_row, data_start, data_start + rows - 1, labels, [field.column_width for field in fields])
        renderer.configure_table(sheet, header_row, data_start + rows - 1, len(fields))

    @staticmethod
    def _formula(expression: str, columns: dict[str, str], row: int) -> str:
        converted = FIELD_REFERENCE.sub(lambda match: f"{columns[match.group(1)]}{row}", expression)
        return f"=IFERROR({converted},0)" if "/" in converted else f"={converted}"

    def _write_system_data(self, sheet, data: dict[str, list[str]]) -> None:
        for column, (_key, values) in enumerate(data.items(), 1):
            for row, value in enumerate(values, 1):
                sheet.cell(row, column, value)

    def _write_instructions(self, sheet, schema: TemplateSchema, renderer: ExcelStyleRenderer) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.append(["模板名称", schema.template_name])
        sheet.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        sheet.append(["说明", schema.description or "无"])
        sheet.append([])
        sheet.append(["字段", "必填", "说明"])
        renderer.style_header(sheet, 5, ["字段", "必填", "说明"])
        for field in schema.sheets[0].fields:
            sheet.append([field.name, "是" if field.required else "否", field.description or ""])
        if schema.student_related:
            sheet.append([])
            sheet.append(["注意事项", "姓名、学号、班级是学生核心身份字段，导入时用于匹配学生库。"])
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 50
        sheet.column_dimensions["C"].width = 45
