"""模板与班级名单共用的工作簿样式 Schema、预设和渲染器。"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
import re

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.utils.excel_safety import safe_excel_value

STYLE_PRESETS = ("标准办公表格", "商务蓝色", "极简表格", "自定义")
COLOR = re.compile(r"^[0-9A-Fa-f]{6}$")


@dataclass(frozen=True)
class WorkbookStyleSchema:
    preset: str = "标准办公表格"
    overall_font_name: str = "宋体"
    overall_font_size: int = 11
    show_main_title: bool = False
    main_title: str = ""
    title_font_name: str = "宋体"
    title_font_size: int = 18
    title_bold: bool = True
    title_horizontal_alignment: str = "center"
    title_vertical_alignment: str = "center"
    title_row_height: float = 30
    header_fill_enabled: bool = False
    header_fill_color: str = "FFFFFF"
    header_font_color: str = "000000"
    header_font_size: int = 11
    header_bold: bool = True
    header_horizontal_alignment: str = "center"
    header_vertical_alignment: str = "center"
    header_row_height: float = 24
    header_wrap_text: bool = True
    body_font_size: int = 11
    body_horizontal_alignment: str = "left"
    body_vertical_alignment: str = "center"
    body_row_height: float = 22
    body_wrap_text: bool = True
    border_enabled: bool = True
    border_color: str = "000000"
    border_style: str = "thin"
    show_gridlines: bool = False
    freeze_header: bool = True
    auto_filter: bool = True
    default_column_width: float | None = None
    required_display: str = "none"
    required_header_color: str = "C00000"
    required_cell_fill_color: str = "FFF2CC"

    @classmethod
    def from_dict(cls, value: dict | None) -> "WorkbookStyleSchema":
        if not value:
            return standard_office_style()
        allowed = {key: item for key, item in value.items() if key in cls.__dataclass_fields__}
        return cls(**allowed)

    def validate(self) -> None:
        if self.preset not in STYLE_PRESETS:
            raise ValueError("样式预设不支持。")
        for name, size in (("整体字号", self.overall_font_size), ("标题字号", self.title_font_size), ("表头字号", self.header_font_size), ("数据字号", self.body_font_size)):
            if not 6 <= size <= 72:
                raise ValueError(f"{name}必须在 6 到 72 之间。")
        for name, height in (("标题行高", self.title_row_height), ("表头行高", self.header_row_height), ("数据行高", self.body_row_height)):
            if not 12 <= height <= 120:
                raise ValueError(f"{name}必须在 12 到 120 之间。")
        if self.default_column_width is not None and not 5 <= self.default_column_width <= 80:
            raise ValueError("默认列宽必须在 5 到 80 之间。")
        if self.border_style not in {"thin", "medium", "thick"}:
            raise ValueError("边框仅支持细、中、粗。")
        for name, color in (("表头填充色", self.header_fill_color), ("表头字体色", self.header_font_color), ("边框色", self.border_color), ("必填表头色", self.required_header_color), ("必填单元格填充色", self.required_cell_fill_color)):
            if not COLOR.fullmatch(color):
                raise ValueError(f"{name}必须是 6 位 RGB 十六进制颜色。")
        if self.required_display not in {"none", "asterisk", "header_color", "cell_fill"}:
            raise ValueError("必填提示方式不支持。")


def standard_office_style() -> WorkbookStyleSchema:
    return WorkbookStyleSchema()


def business_blue_style() -> WorkbookStyleSchema:
    return WorkbookStyleSchema(
        preset="商务蓝色", header_fill_enabled=True, header_fill_color="1F4E78", header_font_color="FFFFFF",
        border_color="D9E2F3", required_display="cell_fill", body_vertical_alignment="top",
    )


def minimal_style() -> WorkbookStyleSchema:
    return WorkbookStyleSchema(preset="极简表格", border_enabled=False, show_gridlines=True, header_bold=True, body_wrap_text=False)


def preset_style(name: str) -> WorkbookStyleSchema:
    return {"商务蓝色": business_blue_style, "极简表格": minimal_style}.get(name, standard_office_style)()


class ExcelStyleRenderer:
    """仅渲染新生成的工作簿，不用于已有模板填充。"""
    def __init__(self, style: WorkbookStyleSchema) -> None:
        style.validate()
        self.style = style

    def prepare_sheet(self, sheet) -> None:
        sheet.sheet_view.showGridLines = self.style.show_gridlines

    def write_title(self, sheet, column_count: int, fallback_title: str) -> int:
        if not self.style.show_main_title:
            return 1
        title = self.style.main_title.strip() or fallback_title
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
        cell = sheet.cell(1, 1, safe_excel_value(title))
        cell.font = Font(name=self.style.title_font_name, size=self.style.title_font_size, bold=self.style.title_bold, color="000000")
        cell.alignment = Alignment(horizontal=self.style.title_horizontal_alignment, vertical=self.style.title_vertical_alignment, wrap_text=True)
        sheet.row_dimensions[1].height = self.style.title_row_height
        return 2

    def style_header(self, sheet, row: int, labels: list[str]) -> None:
        for index, label in enumerate(labels, 1):
            cell = sheet.cell(row, index, label)
            cell.font = Font(name=self.style.overall_font_name, size=self.style.header_font_size, bold=self.style.header_bold, color=self.style.header_font_color)
            if self.style.header_fill_enabled:
                cell.fill = PatternFill("solid", fgColor=self.style.header_fill_color)
            cell.alignment = Alignment(horizontal=self.style.header_horizontal_alignment, vertical=self.style.header_vertical_alignment, wrap_text=self.style.header_wrap_text)
            cell.border = self._border()
        sheet.row_dimensions[row].height = self.style.header_row_height

    def style_body_cell(self, cell) -> None:
        cell.font = Font(name=self.style.overall_font_name, size=self.style.body_font_size, color="000000")
        cell.alignment = Alignment(horizontal=self.style.body_horizontal_alignment, vertical=self.style.body_vertical_alignment, wrap_text=self.style.body_wrap_text)
        cell.border = self._border()

    def mark_required(self, header_cell, body_cells) -> None:
        if self.style.required_display == "asterisk" and not str(header_cell.value).endswith("*"):
            header_cell.value = f"{header_cell.value}*"
        elif self.style.required_display == "header_color":
            font = copy(header_cell.font)
            font.color = self.style.required_header_color
            header_cell.font = font
        elif self.style.required_display == "cell_fill":
            fill = PatternFill("solid", fgColor=self.style.required_cell_fill_color)
            for cell in body_cells:
                cell.fill = fill

    def set_column_width(self, sheet, index: int, label: str, explicit: float | None = None) -> None:
        width = explicit or self.style.default_column_width or min(max(len(label) + 4, 12), 35)
        sheet.column_dimensions[get_column_letter(index)].width = width

    def configure_table(self, sheet, header_row: int, end_row: int, column_count: int) -> None:
        if self.style.freeze_header:
            sheet.freeze_panes = f"A{header_row + 1}"
        if self.style.auto_filter:
            sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(column_count)}{max(header_row, end_row)}"

    def _border(self) -> Border:
        if not self.style.border_enabled:
            return Border()
        side = Side(style=self.style.border_style, color=self.style.border_color)
        return Border(left=side, right=side, top=side, bottom=side)
